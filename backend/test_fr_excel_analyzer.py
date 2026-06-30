from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Side

from app.schemas.agent.fr_report.report_dsl import FieldRole
from app.services.agent.fr_report.excel_analyzer import excel_analyzer


def _sample_net_position_workbook() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "经营净头寸"
    worksheet.merge_cells("A1:J1")
    worksheet["A1"] = "健源公司*年*月*日-*年*月*日经营净头寸"
    worksheet.merge_cells("A2:A3")
    worksheet["A2"] = "项目"
    worksheet.merge_cells("B2:G2")
    worksheet["B2"] = "订单情况"
    worksheet.merge_cells("H2:M2")
    worksheet["H2"] = "原料情况（现货）"
    worksheet.merge_cells("N2:O2")
    worksheet["N2"] = "合同"
    worksheet.merge_cells("P2:P3")
    worksheet["P2"] = "净头寸\n（吨）"
    headers = [
        "期初剩余订单量\n（吨）",
        "核心客户签单量\n（吨）",
        "中小客户签单量\n（吨）",
        "发货量\n（吨）",
        "当日剩余订单量\n（吨）",
        "玉米需求量\n（吨）",
        "期初现货库存\n（吨）",
        "即期采购\n（吨）",
        "合同到货\n（吨）",
        "出库\n（吨）",
        "溢余等\n（吨）",
        "现货结余（吨）",
        "签订\n（吨）",
        "待执行\n（吨）",
    ]
    for column_index, header in enumerate(headers, start=2):
        worksheet.cell(3, column_index).value = header

    worksheet["A4"] = "2026-05-01"
    worksheet["B4"] = 10
    worksheet["C4"] = 2
    worksheet["D4"] = 3
    worksheet["E4"] = 4
    worksheet["F4"] = "=B4+C4+D4-E4"
    worksheet["G4"] = "=F4*1.2"
    worksheet["H4"] = 20
    worksheet["I4"] = 2
    worksheet["J4"] = 3
    worksheet["K4"] = "=E4*1.2"
    worksheet["L4"] = 1
    worksheet["M4"] = "=H4+I4+J4-K4-L4"
    worksheet["N4"] = 7
    worksheet["O4"] = 5
    worksheet["P4"] = "=M4-G4-O4"
    worksheet["A5"] = "2026-05-02"
    worksheet["B5"] = "=F4"
    worksheet["F5"] = "=B5+C5+D5-E5"
    worksheet["G5"] = "=F5*1.2"
    worksheet["H5"] = "=M4"
    worksheet["K5"] = "=E5*1.2"
    worksheet["M5"] = "=H5+I5+J5-K5-L5"
    worksheet["P5"] = "=M5-G5-O5"

    notes = [
        "填报依据：",
        "1、期初剩余订单量确定后手工录入",
        "3、发货量：取自SAP中内贸、外贸发货数据",
        "4、期初现货库存=上期现货结余+即期采购+合同到货-出库",
        "5、原料出库=发货量*1.19",
        "9、净头寸=现货结余+合同待执行-玉米需求量",
        "需求：按时间段筛选",
    ]
    for row_index, note in enumerate(notes, start=13):
        worksheet.cell(row_index, 1).value = note

    worksheet["XFD1"].border = Border(left=Side(style="thin"))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_excel_analyzer_trims_tail_columns_and_extracts_formula_conflicts() -> None:
    result = excel_analyzer.analyze(_sample_net_position_workbook(), "经营净头寸.xlsx")

    sheet = result.sheets[0]
    template = sheet.templateAnalysis or {}

    assert template["effectiveRange"]["address"] == "A1:P19"
    assert len(sheet.fields) == 16
    assert next(field for field in sheet.fields if "当日剩余订单量" in field.label).role == FieldRole.MEASURE

    formula_rules = template["formulaRules"]
    assert {item["targetLabel"].replace("\n", "") for item in formula_rules} >= {
        "出库（吨）",
        "净头寸（吨）",
    }
    conflicts = template["formulaConflicts"]
    assert any(item["type"] == "constant_mismatch" and "出库" in item["targetLabel"] for item in conflicts)
    assert any(item["type"] == "operator_mismatch" and "净头寸" in item["targetLabel"] for item in conflicts)
