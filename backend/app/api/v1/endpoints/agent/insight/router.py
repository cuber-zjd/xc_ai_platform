import asyncio
import json
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from sqlmodel import func, select
from sqlmodel.ext.asyncio.session import AsyncSession

from app.api.deps import get_current_user, get_db
from app.models.agent.insight import InsightChannelAdapterRun
from app.models.system.sys_user import SysUser
from app.schemas.agent.insight.adapter import InsightChannelAdapterDefinitionRead, InsightChannelAdapterRunRead
from app.schemas.agent.insight.crawl import (
    InsightManualUrlCrawlRequest,
    InsightManualUrlCrawlResponse,
    InsightSearchDiscoveryRequest,
    InsightSearchDiscoveryResponse,
)
from app.schemas.agent.insight.asset import (
    InsightAiReviewResponse,
    InsightAssetSearchRequest,
    InsightAssetSearchResponse,
    InsightFormalAssetBackfillRequest,
    InsightFormalAssetBackfillResponse,
    InsightGraphResponse,
)
from app.schemas.agent.insight.channel import InsightChannelCreate, InsightChannelRead, InsightChannelUpdate
from app.schemas.agent.insight.company import (
    InsightCompanyCreate,
    InsightCompanyDetail,
    InsightCompanyImportResponse,
    InsightCompanyListItem,
    InsightCompanyRead,
    InsightCompanyUpdate,
)
from app.schemas.agent.insight.dashboard import InsightDashboardSummary
from app.schemas.agent.insight.data_source import (
    InsightDataSourceBulkActionRequest,
    InsightDataSourceBulkActionResponse,
    InsightDataSourceBatchCreateRequest,
    InsightDataSourceBatchCreateResponse,
    InsightDataSourceCreate,
    InsightDataSourceExecuteRequest,
    InsightDataSourceExecuteResponse,
    InsightDataSourceGroupRead,
    InsightDataSourceImportResponse,
    InsightDataSourceRead,
    InsightDataSourceScheduleRunResponse,
    InsightDataSourceUpdate,
    InsightRequirementSeedRequest,
    InsightRequirementSeedResponse,
    InsightSchedulerStatusRead,
    InsightStaleTaskCleanupResponse,
)
from app.schemas.agent.insight.health import InsightHealthRead
from app.schemas.agent.insight.dictionary import (
    InsightDictionaryOverview,
    InsightIntelligenceTypeRead,
    InsightTagCategoryCreate,
    InsightTagCategoryRead,
    InsightTagCategoryUpdate,
    InsightTagCreate,
    InsightTagRead,
    InsightTagUpdate,
)
from app.schemas.agent.insight.intelligence import (
    InsightCandidatePromoteRequest,
    InsightCandidateReviewRequest,
    InsightCandidateReviewResponse,
    InsightAssistantChatRequest,
    InsightAssistantChatResponse,
    InsightDeepResearchRequest,
    InsightDeepResearchResponse,
    InsightIntelligenceBulkActionRequest,
    InsightIntelligenceBulkActionResponse,
    InsightIntelligenceDetail,
    InsightIntelligenceCandidateListItem,
    InsightIntelligenceCreate,
    InsightIntelligenceListItem,
    InsightIntelligenceSourceCreate,
    InsightIntelligenceSourceRead,
    InsightIntelligenceUpdate,
    InsightPoolUpsertRequest,
    InsightUserIntelligencePoolRead,
    InsightVisibilityRuleCreate,
    InsightVisibilityRuleRead,
)
from app.schemas.agent.insight.report import (
    InsightReportDetail,
    InsightReportExportRead,
    InsightReportExportRequest,
    InsightReportGenerateRequest,
    InsightReportGenerateResponse,
    InsightReportListItem,
    InsightReportPreferenceRead,
    InsightReportPreferenceUpdate,
    InsightReportSubscriptionCreate,
    InsightReportSubscriptionDueRunResponse,
    InsightReportSubscriptionRead,
    InsightReportSubscriptionRunResponse,
    InsightReportSubscriptionUpdate,
    InsightReportTemplateCloneRequest,
    InsightReportTemplateCreate,
    InsightReportTemplatePublishRequest,
    InsightReportTemplateRead,
    InsightReportTemplateUploadResponse,
    InsightReportTemplateUpdate,
    InsightReportUpdateRequest,
)
from app.schemas.agent.insight.permission import InsightAccessRuleBulkResponse, InsightAccessRuleBulkUpsert, InsightAccessRuleRead, InsightAccessRuleUpsert
from app.schemas.agent.insight.notification import InsightNotificationCreate, InsightNotificationRead
from app.schemas.agent.insight.quality import InsightQualityOverview
from app.schemas.agent.insight.settings import InsightSettingsStatusRead
from app.schemas.agent.insight.task import InsightTaskRead
from app.schemas.agent.insight.monitor_config import (
    InsightLegacySourceSyncResponse,
    InsightMonitorConfigCreate,
    InsightMonitorConfigRead,
    InsightMonitorConfigUpdate,
)
from app.schemas.page import Page
from app.schemas.result import Result
from app.services.agent.insight.crawler import insight_crawl_service, insight_search_discovery_service
from app.services.agent.insight.ai_review_service import insight_ai_review_service
from app.services.agent.insight.assistant_service import insight_assistant_service
from app.services.agent.insight.asset_service import insight_asset_service
from app.services.agent.insight.channel_service import insight_channel_service
from app.services.agent.insight.company_service import insight_company_service
from app.services.agent.insight.data_source_service import insight_data_source_service
from app.services.agent.insight.health_service import insight_health_service
from app.services.agent.insight.dictionary_service import insight_dictionary_service
from app.services.agent.insight.intelligence_service import insight_intelligence_service
from app.services.agent.insight.report_service import insight_report_service
from app.services.agent.insight.report_subscription_service import insight_report_subscription_service
from app.services.agent.insight.scheduler_service import insight_scheduler_service
from app.services.agent.insight.permission_service import insight_permission_service
from app.services.agent.insight.notification_service import insight_notification_service
from app.services.agent.insight.quality_service import insight_quality_service
from app.services.agent.insight.requirement_import_service import insight_requirement_import_service
from app.services.agent.insight.settings_service import insight_settings_service
from app.services.agent.insight.monitor_config_service import insight_monitor_config_service
from app.services.agent.insight.monitor_execution_service import insight_monitor_execution_service
from app.services.agent.insight.crawler.channel_adapter_service import insight_channel_adapter_service

router = APIRouter()


@router.get("/health", response_model=Result[InsightHealthRead])
async def get_insight_health(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightHealthRead]:
    """返回 Insight 子平台后端模块装配状态。"""
    _ = current_user
    return Result.success(data=insight_health_service.get_health())


@router.get("/dashboard", response_model=Result[InsightDashboardSummary])
async def get_insight_dashboard(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDashboardSummary]:
    result = await insight_intelligence_service.get_dashboard(
        db,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.get("/settings/status", response_model=Result[InsightSettingsStatusRead])
async def get_insight_settings_status(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightSettingsStatusRead]:
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="仅管理员可查看系统设置")
    return Result.success(data=insight_settings_service.get_status())


@router.get("/settings/channels", response_model=Result[Page[InsightChannelRead]])
async def list_channels(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    channel_type: str | None = None,
    access_status: str | None = None,
    status: str | None = None,
    scenario: str | None = None,
) -> Result[Page[InsightChannelRead]]:
    result = await insight_channel_service.list_channels(
        db,
        page=page,
        size=size,
        keyword=keyword,
        channel_type=channel_type,
        access_status=access_status,
        status=status,
        scenario=scenario,
    )
    return Result.success(data=result)


@router.post("/settings/channels", response_model=Result[InsightChannelRead])
async def create_channel(
    *,
    payload: InsightChannelCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightChannelRead]:
    _ensure_admin(current_user, "仅管理员可维护渠道库")
    try:
        result = await insight_channel_service.create_channel(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="渠道已创建")


@router.put("/settings/channels/{channel_id}", response_model=Result[InsightChannelRead])
async def update_channel(
    *,
    channel_id: int,
    payload: InsightChannelUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightChannelRead]:
    _ensure_admin(current_user, "仅管理员可维护渠道库")
    try:
        result = await insight_channel_service.update_channel(db, channel_id, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="渠道已更新")


@router.delete("/settings/channels/{channel_id}", response_model=Result[None])
async def delete_channel(
    *,
    channel_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    _ensure_admin(current_user, "仅管理员可维护渠道库")
    try:
        await insight_channel_service.delete_channel(db, channel_id, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(msg="渠道已删除")


@router.post("/settings/channels/seed-defaults", response_model=Result[dict[str, int]])
async def seed_default_channels(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[dict[str, int]]:
    _ensure_admin(current_user, "仅管理员可维护渠道库")
    result = await insight_channel_service.seed_default_channels(db, current_user.id)
    return Result.success(data=result, msg="默认渠道已补齐")


@router.get("/settings/channels/adapters", response_model=Result[list[InsightChannelAdapterDefinitionRead]])
async def list_channel_adapters(
    *,
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightChannelAdapterDefinitionRead]]:
    """查询已迁移的渠道适配器注册状态。"""
    _ensure_admin(current_user, "仅管理员可查看渠道适配器")
    rows = [InsightChannelAdapterDefinitionRead(**item) for item in insight_channel_adapter_service.definitions()]
    return Result.success(data=rows)


@router.get("/quality/adapter-runs", response_model=Result[Page[InsightChannelAdapterRunRead]])
async def list_adapter_runs(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    channel_code: str | None = None,
    status: str | None = None,
    run_type: str | None = None,
) -> Result[Page[InsightChannelAdapterRunRead]]:
    """查询渠道适配器运行、失败和重试审计记录。"""
    _ensure_admin(current_user, "仅管理员可查看渠道运行审计")
    page = max(page, 1)
    size = min(max(size, 1), 100)
    filters = [InsightChannelAdapterRun.is_deleted == 0]
    if channel_code:
        filters.append(InsightChannelAdapterRun.channel_code == channel_code)
    if status:
        filters.append(InsightChannelAdapterRun.status == status)
    if run_type:
        filters.append(InsightChannelAdapterRun.run_type == run_type)
    total = (await db.exec(select(func.count()).select_from(InsightChannelAdapterRun).where(*filters))).one()
    rows = list(
        (
            await db.exec(
                select(InsightChannelAdapterRun)
                .where(*filters)
                .order_by(InsightChannelAdapterRun.started_at.desc())
                .offset((page - 1) * size)
                .limit(size)
            )
        ).all()
    )
    return Result.success(
        data=Page.create(
            items=[InsightChannelAdapterRunRead.model_validate(row) for row in rows],
            total=total,
            page=page,
            size=size,
        )
    )


@router.get("/settings/monitor-configs", response_model=Result[Page[InsightMonitorConfigRead]])
async def list_monitor_configs(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    monitor_type: str | None = None,
    status: str | None = None,
) -> Result[Page[InsightMonitorConfigRead]]:
    result = await insight_monitor_config_service.list_configs(
        db,
        page=page,
        size=size,
        keyword=keyword,
        monitor_type=monitor_type,
        status=status,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/settings/monitor-configs", response_model=Result[InsightMonitorConfigRead])
async def create_monitor_config(
    *,
    payload: InsightMonitorConfigCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightMonitorConfigRead]:
    try:
        result = await insight_monitor_config_service.create_config(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="监测配置已创建")


@router.put("/settings/monitor-configs/{config_id}", response_model=Result[InsightMonitorConfigRead])
async def update_monitor_config(
    *,
    config_id: int,
    payload: InsightMonitorConfigUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightMonitorConfigRead]:
    try:
        result = await insight_monitor_config_service.update_config(db, config_id, payload, current_user.id, is_admin=_is_admin(current_user))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="监测配置已更新")


@router.delete("/settings/monitor-configs/{config_id}", response_model=Result[None])
async def delete_monitor_config(
    *,
    config_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    try:
        await insight_monitor_config_service.delete_config(db, config_id, current_user.id, is_admin=_is_admin(current_user))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(msg="监测配置已删除")


@router.post("/settings/legacy-sources/sync", response_model=Result[InsightLegacySourceSyncResponse])
async def sync_legacy_data_sources(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightLegacySourceSyncResponse]:
    _ensure_admin(current_user, "仅管理员可同步旧执行源")
    result = await insight_data_source_service.sync_legacy_sources(db, user_id=current_user.id)
    return Result.success(data=result, msg="旧执行源归属已同步")


@router.get("/quality/overview", response_model=Result[InsightQualityOverview])
async def get_quality_overview(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightQualityOverview]:
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="仅管理员可查看质量运营")
    result = await insight_quality_service.get_overview(db)
    return Result.success(data=result)


@router.get("/dictionaries/overview", response_model=Result[InsightDictionaryOverview])
async def get_dictionary_overview(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDictionaryOverview]:
    _ = current_user
    result = await insight_dictionary_service.get_overview(db)
    return Result.success(data=result)


@router.get("/dictionaries/tags", response_model=Result[list[InsightTagRead]])
async def list_dictionary_tags(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    tag_type: str | None = None,
    include_disabled: bool = False,
) -> Result[list[InsightTagRead]]:
    _ = current_user
    result = await insight_dictionary_service.list_tags(db, tag_type=tag_type, include_disabled=include_disabled)
    return Result.success(data=result)


@router.get("/dictionaries/tag-categories", response_model=Result[list[InsightTagCategoryRead]])
async def list_dictionary_tag_categories(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    include_disabled: bool = False,
) -> Result[list[InsightTagCategoryRead]]:
    _ = current_user
    result = await insight_dictionary_service.list_categories(db, include_disabled=include_disabled)
    return Result.success(data=result)


@router.post("/dictionaries/tag-categories", response_model=Result[InsightTagCategoryRead])
async def create_dictionary_tag_category(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    payload: InsightTagCategoryCreate,
) -> Result[InsightTagCategoryRead]:
    try:
        result = await insight_dictionary_service.create_category(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="分类已创建")


@router.put("/dictionaries/tag-categories/{category_id}", response_model=Result[InsightTagCategoryRead])
async def update_dictionary_tag_category(
    *,
    category_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    payload: InsightTagCategoryUpdate,
) -> Result[InsightTagCategoryRead]:
    try:
        result = await insight_dictionary_service.update_category(db, category_id, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="分类已更新")


@router.post("/dictionaries/tag-categories/{category_id}/disable", response_model=Result[InsightTagCategoryRead])
async def disable_dictionary_tag_category(
    *,
    category_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightTagCategoryRead]:
    try:
        result = await insight_dictionary_service.disable_category(db, category_id, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="分类已禁用")


@router.post("/dictionaries/tags", response_model=Result[InsightTagRead])
async def create_dictionary_tag(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    payload: InsightTagCreate,
) -> Result[InsightTagRead]:
    try:
        result = await insight_dictionary_service.create_tag(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="标签已创建")


@router.put("/dictionaries/tags/{tag_id}", response_model=Result[InsightTagRead])
async def update_dictionary_tag(
    *,
    tag_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    payload: InsightTagUpdate,
) -> Result[InsightTagRead]:
    try:
        result = await insight_dictionary_service.update_tag(db, tag_id, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="标签已更新")


@router.post("/dictionaries/tags/{tag_id}/disable", response_model=Result[InsightTagRead])
async def disable_dictionary_tag(
    *,
    tag_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightTagRead]:
    try:
        result = await insight_dictionary_service.disable_tag(db, tag_id, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="标签已禁用")


@router.get("/dictionaries/intelligence-types", response_model=Result[list[InsightIntelligenceTypeRead]])
async def list_dictionary_intelligence_types(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightIntelligenceTypeRead]]:
    _ = current_user
    result = await insight_dictionary_service.list_intelligence_types(db)
    return Result.success(data=result)


@router.get("/notifications", response_model=Result[Page[InsightNotificationRead]])
async def list_notifications(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    target_type: str | None = None,
    target_id: int | None = None,
    channel: str | None = None,
    status: str | None = None,
) -> Result[Page[InsightNotificationRead]]:
    result = await insight_notification_service.list_notifications(
        db,
        page=page,
        size=size,
        target_type=target_type,
        target_id=target_id,
        channel=channel,
        status=status,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/notifications", response_model=Result[InsightNotificationRead])
async def create_notification(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    payload: InsightNotificationCreate,
) -> Result[InsightNotificationRead]:
    try:
        result = await insight_notification_service.create_notification(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result)


@router.post("/notifications/{notification_id}/retry", response_model=Result[InsightNotificationRead])
async def retry_notification(
    *,
    notification_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightNotificationRead]:
    try:
        result = await insight_notification_service.retry_notification(
            db,
            notification_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result)


@router.get("/companies", response_model=Result[Page[InsightCompanyListItem]])
async def list_companies(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    sys_company_id: int | None = None,
    industry: str | None = None,
    monitor_level: str | None = None,
    status: str | None = None,
) -> Result[Page[InsightCompanyListItem]]:
    result = await insight_company_service.list_companies(
        db,
        page=page,
        size=size,
        keyword=keyword,
        sys_company_id=sys_company_id,
        industry=industry,
        monitor_level=monitor_level,
        status=status,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/companies", response_model=Result[InsightCompanyRead])
async def create_company(
    *,
    payload: InsightCompanyCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCompanyRead]:
    try:
        result = await insight_company_service.create_company(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="企业档案已创建")


@router.post("/companies/import", response_model=Result[InsightCompanyImportResponse])
async def import_companies(
    *,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCompanyImportResponse]:
    try:
        file_bytes = await file.read()
        result = await insight_company_service.import_companies_from_excel(
            db,
            file_name=file.filename,
            file_bytes=file_bytes,
            user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="企业档案导入完成")


@router.get("/companies/import-template")
async def download_company_import_template(
    current_user: SysUser = Depends(get_current_user),
) -> StreamingResponse:
    _ = current_user
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，无法生成 Excel 模板") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "企业档案导入"
    headers = ["企业名称", "简称", "所属公司ID", "所属公司", "行业", "企业类型", "区域", "官网", "监控级别", "描述"]
    example = [
        "某客户食品有限公司",
        "某客户食品",
        "",
        "山东香驰健源生物科技有限公司",
        "食品饮料",
        "客户",
        "华东",
        "https://www.example.com",
        "重点客户",
        "饮料和休闲食品客户，重点关注低糖配方、功能配料和新品上市动态",
    ]
    sheet.append(headers)
    sheet.append(example)
    sheet.append([
        "某植物蛋白科技有限公司",
        "某植物蛋白",
        "",
        "山东御馨生物科技股份有限公司",
        "植物蛋白",
        "竞对",
        "华南",
        "",
        "重点竞对",
        "植物蛋白和饮品应用相关企业，重点关注产能、客户合作、产品应用和价格策略",
    ])

    header_fill = PatternFill("solid", fgColor="DDEBFF")
    header_font = Font(bold=True, color="17365D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate([26, 16, 14, 28, 16, 14, 14, 28, 14, 36], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    note_sheet = workbook.create_sheet("填写说明")
    note_sheet.append(["字段", "填写说明"])
    note_sheet.append(["企业名称", "必填。请填写企业工商全称；如果企业编码或企业名称已存在，导入时会更新原档案。"])
    note_sheet.append(["所属公司ID", "可选。已知 sys_company.id 时优先填写 ID。"])
    note_sheet.append(["所属公司", "可选但建议填写。必须与系统公司名称一致，例如：山东香驰健源生物科技有限公司、山东御馨生物科技股份有限公司。"])
    note_sheet.append(["企业类型", "建议填写客户、竞对、供应商、渠道、行业机构等受控口径，便于后续筛选和报告分组。"])
    note_sheet.append(["监控级别", "建议填写重点客户、重点竞对、普通关注等，便于数据源和报告优先级配置。"])
    note_sheet.append(["描述", "建议填写产品线、经营重点、关注原因、合作或竞争关系等有助于 AI 判断的信息，不要填写来源说明等重复信息。"])
    for cell in note_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate([16, 88], start=1):
        note_sheet.column_dimensions[get_column_letter(index)].width = width
    note_sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers_map = {"Content-Disposition": 'attachment; filename="insight-company-import-template.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_map,
    )


@router.get("/companies/{company_id}", response_model=Result[InsightCompanyDetail])
async def get_company_detail(
    *,
    company_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCompanyDetail]:
    try:
        result = await insight_company_service.get_company_detail(
            db,
            company_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result)


@router.put("/companies/{company_id}", response_model=Result[InsightCompanyRead])
async def update_company(
    *,
    company_id: int,
    payload: InsightCompanyUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCompanyRead]:
    try:
        result = await insight_company_service.update_company(
            db,
            company_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="企业档案已更新")


@router.post("/crawler/manual-url", response_model=Result[InsightManualUrlCrawlResponse])
async def crawl_manual_url(
    *,
    payload: InsightManualUrlCrawlRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightManualUrlCrawlResponse]:
    try:
        result = await insight_crawl_service.crawl_manual_url(
            db,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="网页抓取完成")


@router.post("/crawler/search-discovery", response_model=Result[InsightSearchDiscoveryResponse])
async def search_discovery(
    *,
    payload: InsightSearchDiscoveryRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightSearchDiscoveryResponse]:
    try:
        result = await insight_search_discovery_service.search_and_crawl(
            db,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="搜索发现与抓取完成")


@router.get("/reports", response_model=Result[Page[InsightReportListItem]])
async def list_reports(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    report_type: str | None = None,
    status: str | None = None,
) -> Result[Page[InsightReportListItem]]:
    result = await insight_report_service.list_reports(
        db,
        page=page,
        size=size,
        keyword=keyword,
        report_type=report_type,
        status=status,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/reports/generate", response_model=Result[InsightReportGenerateResponse])
async def generate_report(
    *,
    payload: InsightReportGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportGenerateResponse]:
    try:
        result = await insight_report_service.generate_report(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告草稿已生成")


@router.post("/reports/generate/stream")
async def generate_report_stream(
    *,
    payload: InsightReportGenerateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> StreamingResponse:
    queue: asyncio.Queue[dict | None] = asyncio.Queue()

    def encode_event(data: dict) -> str:
        return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"

    async def progress_callback(event: dict) -> None:
        await queue.put(event)

    async def run_generation() -> None:
        try:
            result = await insight_report_service.generate_report(
                db,
                payload,
                user_id=current_user.id,
                is_admin=_is_admin(current_user),
                progress_callback=progress_callback,
            )
            await queue.put(
                {
                    "event": "done",
                    "step": "done",
                    "title": "报告已完成",
                    "detail": "报告草稿已经生成完成。",
                    "progress": 100,
                    "data": result.model_dump(mode="json"),
                }
            )
        except ValueError as exc:
            await queue.put({"event": "error", "title": "生成失败", "detail": str(exc), "progress": 100})
        except Exception:
            await queue.put({"event": "error", "title": "生成失败", "detail": "报告生成过程中出现异常，请稍后重试。", "progress": 100})
        finally:
            await queue.put(None)

    async def event_stream():
        yield encode_event(
            {
                "event": "connected",
                "step": "connected",
                "title": "准备生成报告",
                "detail": "正在启动研究任务。",
                "progress": 1,
            }
        )
        task = asyncio.create_task(run_generation())
        while True:
            event = await queue.get()
            if event is None:
                break
            yield encode_event(event)
        await task

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.get("/reports/subscriptions", response_model=Result[Page[InsightReportSubscriptionRead]])
async def list_report_subscriptions(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    status: str | None = None,
) -> Result[Page[InsightReportSubscriptionRead]]:
    result = await insight_report_subscription_service.list_subscriptions(
        db,
        page=page,
        size=size,
        status=status,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/reports/subscriptions", response_model=Result[InsightReportSubscriptionRead])
async def create_report_subscription(
    *,
    payload: InsightReportSubscriptionCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportSubscriptionRead]:
    try:
        result = await insight_report_subscription_service.create_subscription(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="定时报告计划已保存")


@router.put("/reports/subscriptions/{subscription_id}", response_model=Result[InsightReportSubscriptionRead])
async def update_report_subscription(
    *,
    subscription_id: int,
    payload: InsightReportSubscriptionUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportSubscriptionRead]:
    try:
        result = await insight_report_subscription_service.update_subscription(
            db,
            subscription_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="定时报告计划已更新")


@router.delete("/reports/subscriptions/{subscription_id}", response_model=Result[None])
async def delete_report_subscription(
    *,
    subscription_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    try:
        await insight_report_subscription_service.delete_subscription(
            db,
            subscription_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=None, msg="定时报告计划已删除")


@router.post("/reports/subscriptions/{subscription_id}/run", response_model=Result[InsightReportSubscriptionRunResponse])
async def run_report_subscription(
    *,
    subscription_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportSubscriptionRunResponse]:
    try:
        result = await insight_report_subscription_service.run_subscription(
            db,
            subscription_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
            triggered_by=f"user:{current_user.id}",
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="定时报告计划已执行")


@router.post("/reports/subscriptions/run-due", response_model=Result[InsightReportSubscriptionDueRunResponse])
async def run_due_report_subscriptions(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    limit: int = 10,
) -> Result[InsightReportSubscriptionDueRunResponse]:
    _ = current_user
    result = await insight_report_subscription_service.run_due_subscriptions(db, limit=limit, triggered_by="manual_due_scan")
    return Result.success(data=result, msg="到期定时报告已扫描")


@router.get("/reports/templates", response_model=Result[list[InsightReportTemplateRead]])
async def list_report_templates(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightReportTemplateRead]]:
    result = await insight_report_service.list_templates(
        db,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/reports/templates", response_model=Result[InsightReportTemplateRead])
async def create_report_template(
    *,
    payload: InsightReportTemplateCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportTemplateRead]:
    result = await insight_report_service.create_template(db, payload, user_id=current_user.id)
    return Result.success(data=result, msg="报告模板已保存")


@router.post("/reports/templates/upload", response_model=Result[InsightReportTemplateUploadResponse])
async def upload_report_template(
    *,
    file: UploadFile = File(...),
    template_name: str | None = Form(default=None),
    report_type: str = Form(default="专题报告"),
    description: str | None = Form(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportTemplateUploadResponse]:
    try:
        content = await file.read()
        result = await insight_report_service.create_template_from_upload(
            db,
            file_name=file.filename or "未命名模板",
            file_bytes=content,
            template_name=template_name,
            report_type=report_type,
            description=description,
            user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告模板已解析并保存")


@router.post("/reports/templates/{template_id}/publish", response_model=Result[InsightReportTemplateRead])
async def publish_report_template(
    *,
    template_id: int,
    payload: InsightReportTemplatePublishRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportTemplateRead]:
    try:
        result = await insight_report_service.publish_template(
            db,
            template_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告模板已发布到模板市场")


@router.post("/reports/templates/{template_code}/clone", response_model=Result[InsightReportTemplateRead])
async def clone_report_template(
    *,
    template_code: str,
    payload: InsightReportTemplateCloneRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportTemplateRead]:
    try:
        result = await insight_report_service.clone_template(
            db,
            template_code,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="模板已复制为我的模板")


@router.put("/reports/templates/{template_id}", response_model=Result[InsightReportTemplateRead])
async def update_report_template(
    *,
    template_id: int,
    payload: InsightReportTemplateUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportTemplateRead]:
    try:
        result = await insight_report_service.update_template(
            db,
            template_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告模板已更新")


@router.delete("/reports/templates/{template_id}", response_model=Result[None])
async def delete_report_template(
    *,
    template_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    try:
        await insight_report_service.delete_template(
            db,
            template_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(msg="报告模板已删除")


@router.get("/reports/preference", response_model=Result[InsightReportPreferenceRead])
async def get_report_preference(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportPreferenceRead]:
    result = await insight_report_service.get_preference(db, user_id=current_user.id)
    return Result.success(data=result)


@router.put("/reports/preference", response_model=Result[InsightReportPreferenceRead])
async def update_report_preference(
    *,
    payload: InsightReportPreferenceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportPreferenceRead]:
    result = await insight_report_service.update_preference(db, payload, user_id=current_user.id)
    return Result.success(data=result, msg="报告偏好已保存")


@router.get("/reports/{report_id}", response_model=Result[InsightReportDetail])
async def get_report_detail(
    *,
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportDetail]:
    try:
        result = await insight_report_service.get_report_detail(
            db,
            report_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result)


@router.put("/reports/{report_id}", response_model=Result[InsightReportDetail])
async def update_report(
    *,
    report_id: int,
    payload: InsightReportUpdateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportDetail]:
    try:
        result = await insight_report_service.update_report(
            db,
            report_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告草稿已更新")


@router.get("/reports/{report_id}/exports", response_model=Result[list[InsightReportExportRead]])
async def list_report_exports(
    *,
    report_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightReportExportRead]]:
    try:
        result = await insight_report_service.list_report_exports(
            db,
            report_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result)


@router.post("/reports/{report_id}/exports", response_model=Result[InsightReportExportRead])
async def export_report(
    *,
    report_id: int,
    payload: InsightReportExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightReportExportRead]:
    try:
        result = await insight_report_service.export_report(
            db,
            report_id,
            export_format=payload.export_format,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="报告导出已生成")


@router.get("/reports/{report_id}/exports/{export_id}/download")
async def download_report_export(
    *,
    report_id: int,
    export_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> FileResponse:
    try:
        file_path, export = await insight_report_service.get_report_export_file(
            db,
            report_id,
            export_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return FileResponse(
        path=file_path,
        filename=export.file_name or file_path.name,
        media_type=export.content_type or "application/octet-stream",
    )


@router.get("/permissions/{target_type}/{target_id}", response_model=Result[list[InsightAccessRuleRead]])
async def list_access_rules(
    *,
    target_type: str,
    target_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightAccessRuleRead]]:
    _ = current_user
    result = await insight_permission_service.list_rules(db, target_type=target_type, target_id=target_id)
    return Result.success(data=result)


@router.post("/permissions/{target_type}/bulk", response_model=Result[InsightAccessRuleBulkResponse])
async def grant_access_rules_bulk(
    *,
    target_type: str,
    payload: InsightAccessRuleBulkUpsert,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightAccessRuleBulkResponse]:
    result = await insight_permission_service.grant_rules_bulk(
        db,
        target_type=target_type,
        payload=payload,
        user_id=current_user.id,
    )
    return Result.success(data=result, msg="批量权限已更新")


@router.post("/permissions/{target_type}/{target_id}", response_model=Result[InsightAccessRuleRead])
async def grant_access_rule(
    *,
    target_type: str,
    target_id: int,
    payload: InsightAccessRuleUpsert,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightAccessRuleRead]:
    result = await insight_permission_service.grant_rule(
        db,
        target_type=target_type,
        target_id=target_id,
        payload=payload,
        user_id=current_user.id,
    )
    return Result.success(data=result, msg="权限已更新")


@router.delete("/permissions/rules/{rule_id}", response_model=Result[None])
async def revoke_access_rule(
    *,
    rule_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    try:
        await insight_permission_service.revoke_rule(db, rule_id=rule_id, user_id=current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(msg="权限已撤销")


@router.get("/data-sources", response_model=Result[Page[InsightDataSourceRead]])
async def list_data_sources(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    source_type: str | None = None,
    status: str | None = None,
    monitor_config_id: int | None = None,
    execution_role: str | None = None,
    channel_id: int | None = None,
) -> Result[Page[InsightDataSourceRead]]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    result = await insight_data_source_service.list_data_sources(
        db,
        page=page,
        size=size,
        keyword=keyword,
        source_type=source_type,
        status=status,
        monitor_config_id=monitor_config_id,
        execution_role=execution_role,
        channel_id=channel_id,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.get("/data-sources/groups", response_model=Result[list[InsightDataSourceGroupRead]])
async def list_data_source_groups(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    keyword: str | None = None,
    source_type: str | None = None,
    status: str | None = None,
    monitor_config_id: int | None = None,
    execution_role: str | None = None,
    channel_id: int | None = None,
) -> Result[list[InsightDataSourceGroupRead]]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    result = await insight_data_source_service.list_data_source_groups(
        db,
        keyword=keyword,
        source_type=source_type,
        status=status,
        monitor_config_id=monitor_config_id,
        execution_role=execution_role,
        channel_id=channel_id,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/data-sources", response_model=Result[InsightDataSourceRead])
async def create_data_source(
    *,
    payload: InsightDataSourceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceRead]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        result = await insight_data_source_service.create_data_source(db, payload, current_user.id, is_admin=_is_admin(current_user))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源已创建")


@router.post("/data-sources/batch-create", response_model=Result[InsightDataSourceBatchCreateResponse])
async def batch_create_data_sources(
    *,
    payload: InsightDataSourceBatchCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceBatchCreateResponse]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        result = await insight_data_source_service.batch_create_data_sources(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="标准数据源已批量生成")


@router.get("/data-sources/import-template")
async def download_data_source_import_template(
    current_user: SysUser = Depends(get_current_user),
) -> StreamingResponse:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，无法生成 Excel 模板") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "数据源导入"
    headers = [
        "来源名称",
        "来源类型",
        "官网地址",
        "关键词",
        "包含词",
        "排除词",
        "课题/项目",
        "搜集需求",
        "获取内容",
        "采集周期",
        "所属公司",
        "关联企业",
        "可见范围",
        "自动审核",
        "自动加入素材池",
    ]
    sheet.append(headers)
    sheet.append([
        "Foodaily每日食品",
        "industry_media",
        "",
        "大豆蛋白、植物基、饮料新品",
        "新品、配料、应用",
        "招聘、广告招商",
        "植物基蛋白趋势",
        "跟踪食品饮料新品和应用方案",
        "新品、配方、宣称、渠道",
        "daily",
        "山东御馨生物科技股份有限公司",
        "",
        "assigned",
        "high_confidence",
        "是",
    ])
    header_fill = PatternFill("solid", fgColor="DDEBFF")
    header_font = Font(bold=True, color="17365D")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for index, width in enumerate([24, 20, 34, 34, 26, 24, 24, 40, 34, 14, 34, 24, 14, 16, 18], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    headers_map = {"Content-Disposition": 'attachment; filename="insight-data-source-import-template.xlsx"'}
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_map,
    )


@router.post("/data-sources/import", response_model=Result[InsightDataSourceImportResponse])
async def import_data_sources(
    *,
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceImportResponse]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        payload_files = [(file.filename or "未命名文件", await file.read()) for file in files]
        result = await insight_requirement_import_service.import_files(
            db,
            files=payload_files,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源导入完成")


@router.post("/data-sources/import-preview", response_model=Result[InsightDataSourceImportResponse])
async def preview_import_data_sources(
    *,
    files: list[UploadFile] = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceImportResponse]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        payload_files = [(file.filename or "未命名文件", await file.read()) for file in files]
        result = await insight_requirement_import_service.preview_files(
            db,
            files=payload_files,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源导入预览完成")


@router.post("/data-sources/bulk-action", response_model=Result[InsightDataSourceBulkActionResponse])
async def bulk_action_data_sources(
    *,
    payload: InsightDataSourceBulkActionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceBulkActionResponse]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        result = await insight_data_source_service.bulk_action(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="批量数据源操作完成")


@router.post("/bootstrap/seed-from-requirements", response_model=Result[InsightRequirementSeedResponse])
async def seed_from_requirements(
    *,
    payload: InsightRequirementSeedRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightRequirementSeedResponse]:
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="仅管理员可执行期初数据源导入与采集")
    try:
        local_files: list[tuple[str, bytes]] = []
        for file_path in payload.file_paths:
            with open(file_path, "rb") as file:
                local_files.append((file_path, file.read()))
        import_result = await insight_requirement_import_service.import_files(
            db,
            files=local_files,
            user_id=current_user.id,
            is_admin=True,
        )
        execution_result = None
        if payload.execute:
            ids = [item.data_source_id for item in import_result.items if item.data_source_id][: payload.max_sources_to_execute]
            if ids:
                execution_result = await insight_data_source_service.bulk_action(
                    db,
                    InsightDataSourceBulkActionRequest(
                        data_source_ids=ids,
                        action="execute",
                        execute_crawl_top_n=payload.crawl_top_n,
                    ),
                    user_id=current_user.id,
                    is_admin=True,
                )
        result = InsightRequirementSeedResponse(
            import_result=import_result,
            execution_result=execution_result,
            target_intelligence_count=payload.target_intelligence_count,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="期初需求渠道导入已完成")


@router.get("/data-sources/execution-logs", response_model=Result[Page[InsightTaskRead]])
async def list_data_source_execution_logs(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    data_source_id: int | None = None,
    status: str | None = None,
    task_type: str | None = None,
) -> Result[Page[InsightTaskRead]]:
    _ensure_admin(current_user, "仅管理员可查看执行源日志")
    result = await insight_data_source_service.list_execution_logs(
        db,
        page=page,
        size=size,
        data_source_id=data_source_id,
        status=status,
        task_type=task_type,
    )
    return Result.success(data=result)


@router.post("/data-sources/tasks/cleanup-stale", response_model=Result[InsightStaleTaskCleanupResponse])
async def cleanup_stale_data_source_tasks(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    timeout_minutes: int = 30,
) -> Result[InsightStaleTaskCleanupResponse]:
    _ensure_admin(current_user, "仅管理员可清理执行源任务")
    result = await insight_data_source_service.cleanup_stale_tasks(
        db,
        timeout_minutes=timeout_minutes,
        user_id=current_user.id,
    )
    return Result.success(data=result, msg="遗留任务清理完成")


@router.post("/data-sources/{data_source_id}/schedule/retry", response_model=Result[InsightDataSourceRead])
async def retry_data_source_schedule(
    *,
    data_source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceRead]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        result = await insight_data_source_service.retry_data_source(
            db,
            data_source_id=data_source_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源已加入下一轮调度")


@router.get("/scheduler/status", response_model=Result[InsightSchedulerStatusRead])
async def get_scheduler_status(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightSchedulerStatusRead]:
    _ensure_admin(current_user, "仅管理员可查看调度器状态")
    return Result.success(data=InsightSchedulerStatusRead(**insight_scheduler_service.status()))


@router.post("/scheduler/run-once", response_model=Result[InsightDataSourceScheduleRunResponse])
async def run_scheduler_once(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceScheduleRunResponse]:
    _ensure_admin(current_user, "仅管理员可触发调度器")
    result = await insight_scheduler_service.run_once(triggered_by=f"user:{current_user.id}")
    return Result.success(data=result, msg="调度器已完成一次扫描")


@router.post("/scheduler/start", response_model=Result[InsightSchedulerStatusRead])
async def start_scheduler(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightSchedulerStatusRead]:
    _ensure_admin(current_user, "仅管理员可启动调度器")
    await insight_scheduler_service.start()
    return Result.success(data=InsightSchedulerStatusRead(**insight_scheduler_service.status()), msg="调度器已启动")


@router.post("/scheduler/stop", response_model=Result[InsightSchedulerStatusRead])
async def stop_scheduler(
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightSchedulerStatusRead]:
    _ensure_admin(current_user, "仅管理员可停止调度器")
    await insight_scheduler_service.stop()
    return Result.success(data=InsightSchedulerStatusRead(**insight_scheduler_service.status()), msg="调度器已停止")


@router.post("/data-sources/schedule/run-due", response_model=Result[InsightDataSourceScheduleRunResponse])
async def run_due_data_sources(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    limit: int = 5,
) -> Result[InsightDataSourceScheduleRunResponse]:
    _ensure_admin(current_user, "仅管理员可执行到期监测配置")
    result = await insight_monitor_execution_service.run_due_monitor_configs(
        db,
        limit=limit,
        user_id=current_user.id,
    )
    return Result.success(data=result, msg="到期监测配置采集完成")


@router.get("/data-sources/{data_source_id}", response_model=Result[InsightDataSourceRead])
async def get_data_source(
    *,
    data_source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceRead]:
    _ensure_admin(current_user, "仅管理员可查看执行源")
    try:
        result = await insight_data_source_service.get_data_source(
            db,
            data_source_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result)


@router.put("/data-sources/{data_source_id}", response_model=Result[InsightDataSourceRead])
async def update_data_source(
    *,
    data_source_id: int,
    payload: InsightDataSourceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceRead]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        result = await insight_data_source_service.update_data_source(
            db,
            data_source_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源已更新")


@router.delete("/data-sources/{data_source_id}", response_model=Result[None])
async def delete_data_source(
    *,
    data_source_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    _ensure_admin(current_user, "仅管理员可维护执行源")
    try:
        await insight_data_source_service.delete_data_source(
            db,
            data_source_id,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(msg="数据源已删除")


@router.post("/data-sources/{data_source_id}/execute", response_model=Result[InsightDataSourceExecuteResponse])
async def execute_data_source(
    *,
    data_source_id: int,
    payload: InsightDataSourceExecuteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDataSourceExecuteResponse]:
    _ensure_admin(current_user, "仅管理员可测试执行源")
    try:
        result = await insight_data_source_service.execute_data_source(
            db,
            data_source_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="数据源测试采集已完成")


@router.get("/intelligence/candidates", response_model=Result[Page[InsightIntelligenceCandidateListItem]])
async def list_intelligence_candidates(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    review_status: str | None = None,
    subject_type: str | None = None,
    intelligence_type: str | None = None,
    data_source_id: int | None = None,
) -> Result[Page[InsightIntelligenceCandidateListItem]]:
    result = await insight_intelligence_service.list_candidates(
        db,
        page=page,
        size=size,
        keyword=keyword,
        review_status=review_status,
        subject_type=subject_type,
        intelligence_type=intelligence_type,
        data_source_id=data_source_id,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.get("/intelligence", response_model=Result[Page[InsightIntelligenceListItem]])
async def list_intelligences(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    page: int = 1,
    size: int = 20,
    keyword: str | None = None,
    subject_type: str | None = None,
    intelligence_type: str | None = None,
    visibility_scope: str | None = None,
    company_id: int | None = None,
    sys_company_id: int | None = None,
    project_name: str | None = None,
    sentiment: str | None = None,
    tag: str | None = None,
    data_source_id: int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Result[Page[InsightIntelligenceListItem]]:
    parsed_date_from = _parse_datetime_param(date_from)
    parsed_date_to = _parse_datetime_param(date_to)
    result = await insight_intelligence_service.list_intelligences(
        db,
        page=page,
        size=size,
        keyword=keyword,
        subject_type=subject_type,
        intelligence_type=intelligence_type,
        visibility_scope=visibility_scope,
        company_id=company_id,
        sys_company_id=sys_company_id,
        project_name=project_name,
        sentiment=sentiment,
        tag=tag,
        data_source_id=data_source_id,
        date_from=parsed_date_from,
        date_to=parsed_date_to,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/intelligence/bulk-action", response_model=Result[InsightIntelligenceBulkActionResponse])
async def bulk_action_intelligence(
    *,
    payload: InsightIntelligenceBulkActionRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightIntelligenceBulkActionResponse]:
    try:
        result = await insight_intelligence_service.bulk_action(
            db,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="批量情报操作完成")


@router.post("/assistant/chat", response_model=Result[InsightAssistantChatResponse])
async def insight_assistant_chat(
    *,
    payload: InsightAssistantChatRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightAssistantChatResponse]:
    result = await insight_assistant_service.chat(
        db,
        payload,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/assets/search", response_model=Result[InsightAssetSearchResponse])
async def search_insight_assets(
    *,
    payload: InsightAssetSearchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightAssetSearchResponse]:
    result = await insight_asset_service.search_assets(
        db,
        payload,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/assets/backfill-formal", response_model=Result[InsightFormalAssetBackfillResponse])
async def backfill_formal_insight_assets(
    *,
    payload: InsightFormalAssetBackfillRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightFormalAssetBackfillResponse]:
    if not _is_admin(current_user):
        raise HTTPException(status_code=403, detail="仅管理员可执行正式情报资产回填")
    result = await insight_asset_service.backfill_formal_assets(
        db,
        payload,
        user_id=current_user.id,
    )
    return Result.success(data=result, msg="正式情报资产回填完成")


@router.get("/assets/graph", response_model=Result[InsightGraphResponse])
async def get_insight_asset_graph(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    company_id: int | None = None,
    asset_id: int | None = None,
    limit: int = 80,
) -> Result[InsightGraphResponse]:
    _ = current_user
    result = await insight_asset_service.graph(
        db,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
        company_id=company_id,
        asset_id=asset_id,
        limit=min(max(limit, 1), 200),
    )
    return Result.success(data=result)


@router.post("/research/deep", response_model=Result[InsightDeepResearchResponse])
async def insight_deep_research(
    *,
    payload: InsightDeepResearchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightDeepResearchResponse]:
    result = await insight_assistant_service.deep_research(
        db,
        payload,
        user_id=current_user.id,
        is_admin=_is_admin(current_user),
    )
    return Result.success(data=result)


@router.post("/intelligence", response_model=Result[InsightIntelligenceDetail])
async def create_intelligence(
    *,
    payload: InsightIntelligenceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightIntelligenceDetail]:
    try:
        result = await insight_intelligence_service.create_intelligence(db, payload, current_user.id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="正式情报已新增")


@router.get("/intelligence/{intelligence_id}", response_model=Result[InsightIntelligenceDetail])
async def get_intelligence_detail(
    *,
    intelligence_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightIntelligenceDetail]:
    try:
        result = await insight_intelligence_service.get_intelligence_detail(
            db,
            intelligence_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return Result.success(data=result)


@router.put("/intelligence/{intelligence_id}", response_model=Result[InsightIntelligenceDetail])
async def update_intelligence(
    *,
    intelligence_id: int,
    payload: InsightIntelligenceUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightIntelligenceDetail]:
    try:
        result = await insight_intelligence_service.update_intelligence(
            db,
            intelligence_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="正式情报已更新")


@router.post("/intelligence/{intelligence_id}/sources", response_model=Result[InsightIntelligenceSourceRead])
async def add_intelligence_source(
    *,
    intelligence_id: int,
    payload: InsightIntelligenceSourceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightIntelligenceSourceRead]:
    try:
        result = await insight_intelligence_service.add_source(
            db,
            intelligence_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="来源证据已补充")


@router.get("/intelligence/{intelligence_id}/visibility-rules", response_model=Result[list[InsightVisibilityRuleRead]])
async def list_intelligence_visibility_rules(
    *,
    intelligence_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[list[InsightVisibilityRuleRead]]:
    try:
        result = await insight_intelligence_service.list_visibility_rules(
            db,
            intelligence_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result)


@router.post("/intelligence/{intelligence_id}/visibility-rules", response_model=Result[InsightVisibilityRuleRead])
async def grant_intelligence_visibility(
    *,
    intelligence_id: int,
    payload: InsightVisibilityRuleCreate,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightVisibilityRuleRead]:
    try:
        result = await insight_intelligence_service.grant_visibility(
            db,
            intelligence_id,
            payload,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="情报可见性已授权")


@router.get("/intelligence-pool", response_model=Result[list[InsightUserIntelligencePoolRead]])
async def list_my_intelligence_pool(
    *,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
    pool_type: str | None = None,
) -> Result[list[InsightUserIntelligencePoolRead]]:
    result = await insight_intelligence_service.list_user_pool(db, user_id=current_user.id, pool_type=pool_type)
    return Result.success(data=result)


@router.post("/intelligence/{intelligence_id}/pool", response_model=Result[InsightUserIntelligencePoolRead])
async def upsert_my_intelligence_pool(
    *,
    intelligence_id: int,
    payload: InsightPoolUpsertRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightUserIntelligencePoolRead]:
    try:
        result = await insight_intelligence_service.upsert_user_pool(
            db,
            intelligence_id,
            payload,
            user_id=current_user.id,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="情报池已更新")


@router.delete("/intelligence/{intelligence_id}/pool/{pool_type}", response_model=Result[None])
async def remove_my_intelligence_pool(
    *,
    intelligence_id: int,
    pool_type: str,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[None]:
    await insight_intelligence_service.remove_user_pool(
        db,
        intelligence_id,
        pool_type,
        user_id=current_user.id,
    )
    return Result.success(msg="情报池记录已移除")


@router.post(
    "/intelligence/candidates/{candidate_id}/promote",
    response_model=Result[InsightCandidateReviewResponse],
)
async def promote_intelligence_candidate(
    *,
    candidate_id: int,
    payload: InsightCandidatePromoteRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCandidateReviewResponse]:
    try:
        result = await insight_intelligence_service.promote_candidate(
            db,
            candidate_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="候选情报已转为正式情报")


@router.post(
    "/intelligence/candidates/{candidate_id}/ai-review",
    response_model=Result[InsightAiReviewResponse],
)
async def ai_review_intelligence_candidate(
    *,
    candidate_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightAiReviewResponse]:
    try:
        result = await insight_ai_review_service.review_candidate(
            db,
            candidate_id,
            user_id=current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="AI 自动评审已完成")


@router.post(
    "/intelligence/candidates/{candidate_id}/reject",
    response_model=Result[InsightCandidateReviewResponse],
)
async def reject_intelligence_candidate(
    *,
    candidate_id: int,
    payload: InsightCandidateReviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCandidateReviewResponse]:
    try:
        result = await insight_intelligence_service.reject_candidate(
            db,
            candidate_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="候选情报已驳回")


@router.post(
    "/intelligence/candidates/{candidate_id}/ignore",
    response_model=Result[InsightCandidateReviewResponse],
)
async def ignore_intelligence_candidate(
    *,
    candidate_id: int,
    payload: InsightCandidateReviewRequest,
    db: AsyncSession = Depends(get_db),
    current_user: SysUser = Depends(get_current_user),
) -> Result[InsightCandidateReviewResponse]:
    try:
        result = await insight_intelligence_service.ignore_candidate(
            db,
            candidate_id,
            payload,
            current_user.id,
            is_admin=_is_admin(current_user),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return Result.success(data=result, msg="候选情报已忽略")


def _is_admin(user: SysUser) -> bool:
    return bool(getattr(user, "is_superuser", False) or getattr(user, "role", None) == "admin")


def _ensure_admin(user: SysUser, message: str = "仅管理员可操作") -> None:
    if not _is_admin(user):
        raise HTTPException(status_code=403, detail=message)


def _parse_datetime_param(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        if len(value) == 10:
            return datetime.fromisoformat(f"{value}T00:00:00")
        return datetime.fromisoformat(value)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"日期参数格式不正确：{value}") from exc
