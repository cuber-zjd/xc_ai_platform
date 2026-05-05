from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from app.schemas.fr_ai_report.ai_report import (
    ExcelAnalysisResult,
    ExcelFieldAnalysis,
    ExcelSheetAnalysis,
)
from app.schemas.fr_ai_report.report_dsl import FieldRole, FieldType


class ExcelAnalyzer:
    def analyze(self, file_content: bytes, file_name: str | None = None) -> ExcelAnalysisResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法读取 Excel 文件") from exc

        workbook = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        sheets: list[ExcelSheetAnalysis] = []

        for worksheet in workbook.worksheets:
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(worksheet.max_row or 1, 80),
                    values_only=True,
                )
            )
            header_index = self._detect_header_row(rows)
            if header_index is None:
                continue

            headers = self._normalize_headers(rows[header_index])
            sample_rows_raw = rows[header_index + 1 : header_index + 31]
            fields = self._analyze_fields(headers, sample_rows_raw)
            sample_rows = self._build_sample_rows(headers, sample_rows_raw[:10])

            sheets.append(
                ExcelSheetAnalysis(
                    sheetName=worksheet.title,
                    headerRowIndex=header_index + 1,
                    rowCount=max((worksheet.max_row or 0) - header_index - 1, 0),
                    fields=fields,
                    sampleRows=sample_rows,
                )
            )

        primary_sheet = max(sheets, key=lambda item: item.rowCount).sheetName if sheets else None
        return ExcelAnalysisResult(fileName=file_name, sheets=sheets, primarySheet=primary_sheet)

    def _detect_header_row(self, rows: list[tuple[Any, ...]]) -> int | None:
        best_index: int | None = None
        best_score = 0
        for index, row in enumerate(rows[:20]):
            non_empty = [cell for cell in row if cell not in (None, "")]
            text_count = sum(isinstance(cell, str) for cell in non_empty)
            score = len(non_empty) + text_count
            if len(non_empty) >= 2 and score > best_score:
                best_index = index
                best_score = score
        return best_index

    def _normalize_headers(self, row: tuple[Any, ...]) -> list[str]:
        names: list[str] = []
        seen: Counter[str] = Counter()
        for index, value in enumerate(row):
            base = str(value).strip() if value not in (None, "") else f"字段{index + 1}"
            base = base.replace("\n", " ").replace("\r", " ")
            seen[base] += 1
            names.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return names

    def _analyze_fields(
        self,
        headers: list[str],
        rows: list[tuple[Any, ...]],
    ) -> list[ExcelFieldAnalysis]:
        fields: list[ExcelFieldAnalysis] = []
        for index, header in enumerate(headers):
            values = [row[index] for row in rows if index < len(row)]
            non_empty_values = [value for value in values if value not in (None, "")]
            field_type = self._infer_type(non_empty_values)
            role = self._infer_role(header, field_type, non_empty_values)
            fields.append(
                ExcelFieldAnalysis(
                    name=self._to_field_name(header, index),
                    label=header,
                    type=field_type,
                    role=role,
                    sampleValues=[self._json_value(value) for value in non_empty_values[:5]],
                    nullRate=round(1 - (len(non_empty_values) / len(values)), 4) if values else 0,
                )
            )
        return fields

    def _infer_type(self, values: list[Any]) -> FieldType:
        if not values:
            return FieldType.STRING
        if all(isinstance(value, bool) for value in values):
            return FieldType.BOOLEAN
        if all(isinstance(value, int) and not isinstance(value, bool) for value in values):
            return FieldType.INTEGER
        if all(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in values):
            return FieldType.DECIMAL
        if all(isinstance(value, datetime) for value in values):
            return FieldType.DATETIME
        if all(isinstance(value, (date, datetime)) for value in values):
            return FieldType.DATE

        parsed_dates = sum(self._looks_like_date(value) for value in values)
        if parsed_dates and parsed_dates / len(values) >= 0.8:
            return FieldType.DATE

        parsed_numbers = sum(self._looks_like_number(value) for value in values)
        if parsed_numbers and parsed_numbers / len(values) >= 0.8:
            return FieldType.DECIMAL
        return FieldType.STRING

    def _infer_role(self, header: str, field_type: FieldType, values: list[Any]) -> FieldRole:
        lowered = header.lower()
        date_keywords = ["date", "time", "日期", "时间", "月份", "年度", "年", "月"]
        measure_keywords = ["金额", "数量", "销量", "收入", "成本", "利润", "单价", "余额", "合计", "sum", "amount", "qty", "price"]
        text_keywords = ["备注", "说明", "描述", "地址", "内容", "comment", "desc"]
        if field_type in {FieldType.DATE, FieldType.DATETIME} or any(key in lowered for key in date_keywords):
            return FieldRole.DATE
        if field_type in {FieldType.INTEGER, FieldType.DECIMAL} and any(key in lowered for key in measure_keywords):
            return FieldRole.MEASURE
        if field_type in {FieldType.INTEGER, FieldType.DECIMAL} and len(set(map(str, values[:30]))) > 15:
            return FieldRole.MEASURE
        if any(key in lowered for key in text_keywords):
            return FieldRole.TEXT
        return FieldRole.DIMENSION

    def _build_sample_rows(self, headers: list[str], rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
        sample_rows: list[dict[str, Any]] = []
        for row in rows:
            if all(value in (None, "") for value in row):
                continue
            sample_rows.append(
                {
                    self._to_field_name(header, index): self._json_value(row[index] if index < len(row) else None)
                    for index, header in enumerate(headers)
                }
            )
        return sample_rows

    def _to_field_name(self, label: str, index: int) -> str:
        normalized = "".join(
            char if char.isascii() and char.isalnum() else "_"
            for char in label.strip()
        ).strip("_").lower()
        return normalized or f"field_{index + 1}"

    def _json_value(self, value: Any) -> Any:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _looks_like_number(self, value: Any) -> bool:
        try:
            float(str(value).replace(",", ""))
            return True
        except (TypeError, ValueError):
            return False

    def _looks_like_date(self, value: Any) -> bool:
        text = str(value)
        return any(separator in text for separator in ["-", "/", "年"]) and any(char.isdigit() for char in text)


excel_analyzer = ExcelAnalyzer()
