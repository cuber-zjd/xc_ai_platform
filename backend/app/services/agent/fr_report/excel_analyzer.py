from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
import re
from typing import Any

from app.schemas.agent.fr_report.ai_report import (
    ExcelAnalysisResult,
    ExcelFieldAnalysis,
    ExcelSheetAnalysis,
)
from app.schemas.agent.fr_report.report_dsl import FieldRole, FieldType


class ExcelAnalyzer:
    def analyze(self, file_content: bytes, file_name: str | None = None) -> ExcelAnalysisResult:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError("缺少 openpyxl 依赖，无法读取 Excel 文件") from exc

        workbook = load_workbook(BytesIO(file_content), read_only=False, data_only=True)
        sheets: list[ExcelSheetAnalysis] = []

        for worksheet in workbook.worksheets:
            rows = list(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(worksheet.max_row or 1, 120),
                    values_only=True,
                )
            )
            table_region = self._detect_table_region(rows)
            if table_region is None:
                continue

            header_index = table_region["headerIndex"]
            data_start_index = table_region["dataStartIndex"]
            header_matrix = self._header_matrix(rows, table_region["headerStartIndex"], header_index)
            header_meta = self._build_header_meta(header_matrix)
            headers = [item["label"] for item in header_meta]
            data_rows_raw = self._data_rows_until_gap(rows, data_start_index)
            sample_rows_raw = data_rows_raw[:30]
            fields = self._analyze_fields(header_meta, sample_rows_raw)
            sample_rows = self._build_sample_rows(headers, sample_rows_raw[:10], fields)
            template_analysis = self._analyze_template(
                worksheet,
                rows,
                table_region,
                header_matrix,
                header_meta,
            )

            sheets.append(
                ExcelSheetAnalysis(
                    sheetName=worksheet.title,
                    headerRowIndex=header_index + 1,
                    rowCount=len(data_rows_raw),
                    fields=fields,
                    sampleRows=sample_rows,
                    templateAnalysis=template_analysis,
                )
            )

        primary_sheet = max(sheets, key=lambda item: item.rowCount).sheetName if sheets else None
        return ExcelAnalysisResult(fileName=file_name, sheets=sheets, primarySheet=primary_sheet)

    def _detect_table_region(self, rows: list[tuple[Any, ...]]) -> dict[str, int] | None:
        data_row_index: int | None = None
        for index, row in enumerate(rows[:60]):
            if self._looks_like_data_row(row):
                data_row_index = index
                break

        if data_row_index is None:
            return None

        header_index = self._find_header_row(rows, data_row_index)
        if header_index is None:
            return None

        header_start_index = header_index
        for index in range(header_index - 1, max(header_index - 3, -1), -1):
            if self._looks_like_header_support_row(rows[index]):
                header_start_index = index
            else:
                break

        return {
            "headerStartIndex": header_start_index,
            "headerIndex": header_index,
            "dataStartIndex": data_row_index,
        }

    def _find_header_row(self, rows: list[tuple[Any, ...]], data_row_index: int) -> int | None:
        best_index: int | None = None
        best_score = -1
        start = max(0, data_row_index - 4)
        for index in range(start, data_row_index):
            row = rows[index]
            non_empty = [cell for cell in row if cell not in (None, "")]
            if len(non_empty) < 2:
                continue
            text_cells = sum(1 for cell in non_empty if self._is_textual_header_cell(cell))
            score = (text_cells * 3) + len(non_empty)
            if score > best_score:
                best_score = score
                best_index = index
        continuation_index = data_row_index - 1
        if best_index is not None and continuation_index > best_index:
            continuation_row = rows[continuation_index]
            if self._looks_like_header_continuation_row(continuation_row):
                return continuation_index
        return best_index

    def _header_matrix(
        self,
        rows: list[tuple[Any, ...]],
        header_start_index: int,
        header_end_index: int,
    ) -> list[list[Any]]:
        matrix: list[list[Any]] = []
        width = max((len(row) for row in rows[header_start_index : header_end_index + 1]), default=0)
        for row in rows[header_start_index : header_end_index + 1]:
            matrix.append([self._json_value(row[index] if index < len(row) else None) for index in range(width)])
        return matrix

    def _build_header_meta(self, header_matrix: list[list[Any]]) -> list[dict[str, Any]]:
        width = max((len(row) for row in header_matrix), default=0)
        meta: list[dict[str, Any]] = []
        for column_index in range(width):
            column_values = [
                row[column_index]
                for row in header_matrix
                if column_index < len(row) and row[column_index] not in (None, "")
            ]
            label = self._compose_header_label(column_values, column_index)
            meta.append(
                {
                    "label": label,
                    "parts": [str(value).strip() for value in column_values if str(value).strip()],
                    "rawParts": column_values,
                    "isMetricColumn": self._is_metric_header_column(column_values),
                }
            )
        return meta

    def _normalize_headers(self, row: tuple[Any, ...]) -> list[str]:
        names: list[str] = []
        seen: Counter[str] = Counter()
        for index, value in enumerate(row):
            base = str(value).strip() if value not in (None, "") else f"字段{index + 1}"
            base = base.replace("\n", " ").replace("\r", " ")
            seen[base] += 1
            names.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
        return names

    def _analyze_fields(
        self,
        header_meta: list[dict[str, Any]],
        rows: list[tuple[Any, ...]],
    ) -> list[ExcelFieldAnalysis]:
        fields: list[ExcelFieldAnalysis] = []
        for index, meta in enumerate(header_meta):
            header = meta["label"]
            values = [row[index] for row in rows if index < len(row)]
            non_empty_values = [value for value in values if value not in (None, "")]
            label = self._normalize_semantic_label(header, meta, non_empty_values)
            field_type = self._infer_type(non_empty_values, meta)
            role = self._infer_role(label, field_type, non_empty_values, meta)
            fields.append(
                ExcelFieldAnalysis(
                    name=self._to_field_name(label, index),
                    label=label,
                    type=field_type,
                    role=role,
                    sampleValues=[self._json_value(value, field_type) for value in non_empty_values[:5]],
                    nullRate=round(1 - (len(non_empty_values) / len(values)), 4) if values else 0,
                )
            )
        return fields

    def _data_rows_until_gap(self, rows: list[tuple[Any, ...]], data_start_index: int) -> list[tuple[Any, ...]]:
        data_rows: list[tuple[Any, ...]] = []
        blank_seen = False
        for row in rows[data_start_index : data_start_index + 120]:
            if all(value in (None, "") for value in row):
                if data_rows:
                    blank_seen = True
                continue
            if blank_seen:
                break
            if not self._looks_like_data_row(row):
                break
            data_rows.append(row)
        return data_rows

    def _infer_type(self, values: list[Any], header_meta: dict[str, Any] | None = None) -> FieldType:
        if not values:
            return FieldType.STRING
        header_text = " ".join(str(part) for part in (header_meta or {}).get("parts", []))
        if self._is_date_like_header(header_text) and self._date_like_ratio(values) >= 0.8:
            return FieldType.DATE
        if self._excel_serial_date_ratio(values) >= 0.8 and self._looks_like_row_date_column(header_text, values):
            return FieldType.DATE
        if all(isinstance(value, bool) for value in values):
            return FieldType.BOOLEAN
        if all(isinstance(value, int) and not isinstance(value, bool) for value in values):
            return FieldType.INTEGER
        if all(isinstance(value, (int, float, Decimal)) and not isinstance(value, bool) for value in values):
            return FieldType.DECIMAL
        if all(isinstance(value, datetime) for value in values):
            return FieldType.DATETIME
        if all(isinstance(value, (date, datetime)) for value in values):
            return FieldType.DATE

        parsed_dates = sum(self._looks_like_date(value) for value in values)
        if parsed_dates and parsed_dates / len(values) >= 0.8:
            return FieldType.DATE

        parsed_numbers = sum(self._looks_like_number(value) for value in values)
        if parsed_numbers and parsed_numbers / len(values) >= 0.8:
            return FieldType.DECIMAL
        return FieldType.STRING

    def _infer_role(
        self,
        header: str,
        field_type: FieldType,
        values: list[Any],
        header_meta: dict[str, Any] | None = None,
    ) -> FieldRole:
        lowered = header.lower()
        date_keywords = ["date", "time", "日期", "时间", "月份", "年度", "年份", "年", "月", "日"]
        measure_keywords = ["金额", "数量", "销量", "收入", "成本", "利润", "单价", "余额", "合计", "均价", "涨跌", "sum", "amount", "qty", "price"]
        text_keywords = ["备注", "说明", "描述", "地址", "内容", "comment", "desc"]
        if field_type in {FieldType.DATE, FieldType.DATETIME} or any(key in lowered for key in date_keywords):
            return FieldRole.DATE
        if header_meta and header_meta.get("isMetricColumn") and field_type in {FieldType.INTEGER, FieldType.DECIMAL}:
            return FieldRole.MEASURE
        if field_type in {FieldType.INTEGER, FieldType.DECIMAL} and any(key in lowered for key in measure_keywords):
            return FieldRole.MEASURE
        if field_type in {FieldType.INTEGER, FieldType.DECIMAL} and len(set(map(str, values[:30]))) > 15:
            return FieldRole.MEASURE
        if any(key in lowered for key in text_keywords):
            return FieldRole.TEXT
        return FieldRole.DIMENSION

    def _build_sample_rows(
        self,
        headers: list[str],
        rows: list[tuple[Any, ...]],
        fields: list[ExcelFieldAnalysis],
    ) -> list[dict[str, Any]]:
        sample_rows: list[dict[str, Any]] = []
        for row in rows:
            if all(value in (None, "") for value in row):
                continue
            sample_rows.append(
                {
                    fields[index].name: self._json_value(
                        row[index] if index < len(row) else None,
                        fields[index].type if index < len(fields) else None,
                    )
                    for index, header in enumerate(headers)
                }
            )
        return sample_rows

    def _analyze_template(
        self,
        worksheet,
        rows: list[tuple[Any, ...]],
        table_region: dict[str, int],
        header_matrix: list[list[Any]],
        header_meta: list[dict[str, Any]],
    ) -> dict[str, Any]:
        non_empty_rows = [
            {
                "rowIndex": index + 1,
                "values": [self._json_value(value) for value in row if value not in (None, "")],
            }
            for index, row in enumerate(rows[: min(len(rows), 30)])
            if any(value not in (None, "") for value in row)
        ]
        merged_ranges = [
            str(cell_range)
            for cell_range in getattr(worksheet, "merged_cells", []).ranges
            if cell_range.min_row <= 30
        ]
        title = self._detect_title(non_empty_rows, rows, table_region, worksheet)
        unit = self._detect_unit(non_empty_rows)
        update_text = self._detect_update_text(non_empty_rows)
        note_texts = self._detect_note_texts(non_empty_rows)
        average_label = self._detect_average_label(header_meta, note_texts)
        filters = self._detect_filter_cells(rows[: min(len(rows), table_region["headerStartIndex"] + 1)])
        header_rows = self._header_rows(rows, table_region["headerStartIndex"], table_region["headerIndex"])
        column_headers = [item["label"] for item in header_meta if item["label"]]
        metric_columns = [item["label"] for item in header_meta if item.get("isMetricColumn")]
        dimension_columns = [item["label"] for item in header_meta if not item.get("isMetricColumn")]
        horizontal_expansion = self._infer_horizontal_expansion(header_meta, note_texts)
        data_rows = self._data_rows_until_gap(rows, table_region["dataStartIndex"])
        date_formats = self._infer_date_formats(header_meta, data_rows[:20])
        calculation_rules = self._infer_calculation_rules(note_texts)
        return {
            "title": title,
            "unit": unit,
            "updateText": update_text,
            "averageLabel": average_label,
            "notes": note_texts,
            "filters": filters,
            "headerRows": header_rows,
            "headerMatrix": header_matrix,
            "columnHeaders": column_headers,
            "mergedRanges": merged_ranges[:30],
            "layoutIntent": self._infer_layout_intent(header_rows, merged_ranges),
            "dataStartRowIndex": table_region["dataStartIndex"] + 1,
            "headerStartRowIndex": table_region["headerStartIndex"] + 1,
            "rowDimensionLabels": self._normalize_row_dimension_labels(dimension_columns)[:6],
            "columnGroupLabels": metric_columns[:20],
            "valueLabels": self._infer_value_labels(column_headers, metric_columns),
            "horizontalExpansion": horizontal_expansion,
            "dateFormatHints": date_formats,
            "calculationRules": calculation_rules,
        }

    def _detect_title(
        self,
        non_empty_rows: list[dict[str, Any]],
        rows: list[tuple[Any, ...]],
        table_region: dict[str, int],
        worksheet,
    ) -> str | None:
        candidates: list[tuple[int, int, str]] = []
        header_start = table_region["headerStartIndex"]
        for row in non_empty_rows:
            row_index = int(row["rowIndex"])
            if row_index > header_start + 1:
                continue
            for value in row["values"]:
                text = self._clean_template_text(value)
                if not text or self._is_non_title_text(text):
                    continue
                score = self._title_score(text, row_index, header_start)
                candidates.append((score, -row_index, text))

        for merged_range in getattr(worksheet, "merged_cells", []).ranges:
            if merged_range.min_row > header_start + 1 or merged_range.min_col > 3:
                continue
            value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            text = self._clean_template_text(value)
            if text and not self._is_non_title_text(text):
                candidates.append((self._title_score(text, merged_range.min_row, header_start) + 8, -merged_range.min_row, text))

        if candidates:
            return sorted(candidates, reverse=True)[0][2]
        for row in rows[:header_start]:
            joined = " ".join(self._clean_template_text(value) for value in row if self._clean_template_text(value))
            if joined and not self._is_non_title_text(joined):
                return joined[:80]
        return None

    def _detect_unit(self, non_empty_rows: list[dict[str, Any]]) -> str | None:
        for row in non_empty_rows[:10]:
            for value in row["values"]:
                text = str(value)
                if "单位" in text:
                    return text.strip()
        return None

    def _detect_update_text(self, non_empty_rows: list[dict[str, Any]]) -> str | None:
        for row in non_empty_rows[:10]:
            for value in row["values"]:
                text = str(value).strip()
                if any(keyword in text for keyword in ["更新时间", "更新日期", "最后一次更新"]):
                    return text
        return None

    def _detect_note_texts(self, non_empty_rows: list[dict[str, Any]]) -> list[str]:
        notes: list[str] = []
        note_keywords = ["显示数据关联", "涨跌", "自动计算", "显示颜色", "价格显示", "备注", "说明"]
        for row in non_empty_rows:
            joined = "\n".join(str(value).strip() for value in row["values"] if str(value).strip())
            if joined and any(keyword in joined for keyword in note_keywords):
                notes.append(joined)
        return notes[:8]

    def _detect_average_label(self, header_meta: list[dict[str, Any]], note_texts: list[str]) -> str | None:
        for item in header_meta:
            label = str(item.get("label") or "")
            if "均价" in label:
                return label
            for part in item.get("parts", []):
                text = str(part)
                if "均价" in text:
                    return text
        for note in note_texts:
            match = re.search(r"[\u4e00-\u9fa5A-Za-z0-9]*均价[\u4e00-\u9fa5A-Za-z0-9]*", note)
            if match:
                return match.group(0)
        return None

    def _detect_filter_cells(self, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
        filters: list[dict[str, Any]] = []
        for row_index, row in enumerate(rows[:8], start=1):
            row_values = [value for value in row if value not in (None, "")]
            joined = " ".join(str(value).strip() for value in row_values)
            if not joined:
                continue
            if "筛选条件" in joined or "返回目录" in joined or any(keyword in joined for keyword in ["时间", "市场", "日照"]):
                filters.append(
                    {
                        "rowIndex": row_index,
                        "values": [self._json_value(value) for value in row_values],
                    }
                )
        return filters

    def _header_rows(self, rows: list[tuple[Any, ...]], header_start_index: int, header_end_index: int) -> list[list[Any]]:
        return [
            [self._json_value(value) for value in row if value not in (None, "")]
            for row in rows[header_start_index : header_end_index + 1]
            if any(value not in (None, "") for value in row)
        ]

    def _column_headers(self, header_rows: list[list[Any]]) -> list[str]:
        headers: list[str] = []
        for row in header_rows:
            for value in row:
                text = str(value).strip()
                if text and text not in headers:
                    headers.append(text)
        return headers

    def _infer_layout_intent(self, header_rows: list[list[Any]], merged_ranges: list[str]) -> str:
        if len(header_rows) >= 2 or merged_ranges:
            return "multi_header_table"
        return "flat_table"

    def _infer_value_labels(self, column_headers: list[str], metric_columns: list[str]) -> list[str]:
        keywords = ["价格", "金额", "数量", "销量", "库存", "涨跌", "合计", "均价", "price", "amount", "qty"]
        inferred = [header for header in column_headers if any(keyword in header.lower() for keyword in keywords)]
        return list(dict.fromkeys(metric_columns + inferred))

    def _infer_horizontal_expansion(self, header_meta: list[dict[str, Any]], note_texts: list[str]) -> dict[str, Any]:
        city_labels = [item["label"] for item in header_meta if item.get("isMetricColumn")]
        value_labels = ["价格"]
        if any("涨跌" in note for note in note_texts) or any("涨跌" in str(part) for item in header_meta for part in item.get("parts", [])):
            value_labels.append("涨跌")
        return {
            "enabled": len(city_labels) >= 3,
            "direction": "right",
            "dimensionLabel": "市场",
            "sourceLabels": city_labels[:30],
            "valueLabels": list(dict.fromkeys(value_labels)),
            "sqlShape": "long_table_preferred",
            "designerIntent": "使用 FineReport 横向扩展按市场展开列，SQL 保持 record_date/market/value 类长表结果。",
        }

    def _infer_date_formats(
        self,
        header_meta: list[dict[str, Any]],
        rows: list[tuple[Any, ...]],
    ) -> list[dict[str, Any]]:
        hints: list[dict[str, Any]] = []
        for index, meta in enumerate(header_meta):
            values = [row[index] for row in rows if index < len(row) and row[index] not in (None, "")]
            label = self._normalize_semantic_label(meta["label"], meta, values)
            if not values:
                continue
            if self._is_date_like_header(label) or self._excel_serial_date_ratio(values) >= 0.8:
                hints.append(
                    {
                        "label": label,
                        "sampleValues": [self._json_value(value, FieldType.DATE) for value in values[:5]],
                        "sourceFormat": "excel_serial_date" if self._excel_serial_date_ratio(values) >= 0.8 else "date_text",
                        "displayFormat": "yyyy年" if "年" in label and "月" not in label and "日" not in label else "MM月dd日",
                    }
                )
        return hints[:6]

    def _infer_calculation_rules(self, note_texts: list[str]) -> list[dict[str, Any]]:
        rules: list[dict[str, Any]] = []
        joined = "\n".join(note_texts)
        if "涨跌" in joined:
            rules.append(
                {
                    "label": "涨跌",
                    "expression": "当日价格 - 前一日价格",
                    "formatIntent": "涨为红色，跌为绿色，不变为黑色",
                    "sqlFieldCandidates": ["change_amt"],
                    "designerHandling": "优先使用数据表已有涨跌字段；缺失时再由 SQL 用 LAG(price) 计算。",
                }
            )
        if "价格显示整数" in joined:
            rules.append({"label": "价格", "formatIntent": "整数显示", "sqlFieldCandidates": ["price"]})
        return rules

    def _normalize_row_dimension_labels(self, labels: list[str]) -> list[str]:
        normalized: list[str] = []
        for label in labels:
            if label == "市场":
                normalized.append("月日")
            else:
                normalized.append(label)
        return list(dict.fromkeys(normalized))

    def _clean_template_text(self, value: Any) -> str:
        if value in (None, ""):
            return ""
        return str(value).replace("\n", " ").replace("\r", " ").strip()

    def _is_non_title_text(self, text: str) -> bool:
        if len(text) <= 1:
            return True
        non_title_keywords = [
            "单位",
            "更新时间",
            "更新日期",
            "筛选",
            "返回目录",
            "备注",
            "说明",
            "市场",
            "日期",
            "年份",
            "涨跌",
        ]
        if any(keyword in text for keyword in non_title_keywords):
            return True
        if self._looks_like_date(text) or self._looks_like_number(text):
            return True
        return False

    def _title_score(self, text: str, row_index: int, header_start_index: int) -> int:
        score = min(len(text), 40)
        if any(keyword in text for keyword in ["报", "表", "汇总", "价格", "行情", "统计", "分析"]):
            score += 20
        if row_index <= 3:
            score += 10
        distance = abs((header_start_index + 1) - row_index)
        return score - distance

    def _compose_header_label(self, values: list[Any], index: int) -> str:
        normalized_parts = [
            str(value).strip()
            for value in values
            if self._is_effective_header_part(value)
        ]
        if normalized_parts:
            unique_parts = list(dict.fromkeys(normalized_parts))
            return unique_parts[0]
        return f"字段{index + 1}"

    def _is_effective_header_part(self, value: Any) -> bool:
        if value in (None, ""):
            return False
        text = str(value).strip()
        if not text:
            return False
        if text in {"0", "-", "--", "/"}:
            return False
        if text == "涨跌":
            return False
        return True

    def _is_metric_header_column(self, values: list[Any]) -> bool:
        parts = [str(value).strip() for value in values if value not in (None, "")]
        if not parts:
            return False
        first = parts[0]
        if any(keyword in first for keyword in ["备注", "市场", "年份", "日期", "时间", "返回目录"]):
            return False
        if len(parts) >= 2 and any(part in {"0", "涨跌"} for part in parts[1:]):
            return True
        if first in {"全国均价"}:
            return True
        return False

    def _normalize_semantic_label(self, header: str, header_meta: dict[str, Any] | None, values: list[Any]) -> str:
        parts = [str(part).strip() for part in (header_meta or {}).get("parts", []) if str(part).strip()]
        if header == "市场" and "涨跌" in parts and self._excel_serial_date_ratio(values) >= 0.8:
            return "月日"
        return header

    def _to_field_name(self, label: str, index: int) -> str:
        normalized = "".join(
            char if char.isascii() and char.isalnum() else "_"
            for char in label.strip()
        ).strip("_").lower()
        return normalized or f"field_{index + 1}"

    def _json_value(self, value: Any, field_type: FieldType | None = None) -> Any:
        if field_type == FieldType.DATE and self._looks_like_excel_date_serial(value):
            return self._excel_serial_to_date(float(value)).isoformat()
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            return float(value)
        return value

    def _looks_like_number(self, value: Any) -> bool:
        try:
            float(str(value).replace(",", ""))
            return True
        except (TypeError, ValueError):
            return False

    def _looks_like_date(self, value: Any) -> bool:
        text = str(value)
        return any(separator in text for separator in ["-", "/", "年", "月", "日"]) and any(char.isdigit() for char in text)

    def _is_date_like_header(self, text: str) -> bool:
        lowered = text.lower()
        if lowered in {"年", "月", "日", "月日"}:
            return True
        return any(keyword in lowered for keyword in ["date", "time", "日期", "时间", "月份", "年度", "年份"])

    def _date_like_ratio(self, values: list[Any]) -> float:
        if not values:
            return 0
        count = sum(self._looks_like_date(value) or isinstance(value, (date, datetime)) for value in values)
        return count / len(values)

    def _excel_serial_date_ratio(self, values: list[Any]) -> float:
        if not values:
            return 0
        count = sum(self._looks_like_excel_date_serial(value) for value in values)
        return count / len(values)

    def _looks_like_row_date_column(self, header_text: str, values: list[Any]) -> bool:
        if self._is_date_like_header(header_text):
            return True
        unique_count = len({str(value) for value in values[:30]})
        return unique_count >= min(3, len(values)) and self._excel_serial_date_ratio(values) >= 0.8

    def _looks_like_excel_date_serial(self, value: Any) -> bool:
        if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
            return False
        serial = float(value)
        return 20000 <= serial <= 80000 and serial.is_integer()

    def _excel_serial_to_date(self, value: float) -> date:
        return date(1899, 12, 30) + timedelta(days=int(value))

    def _looks_like_data_row(self, row: tuple[Any, ...]) -> bool:
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) < 4:
            return False
        if self._looks_like_header_continuation_row(row):
            return False
        numeric_like = sum(
            1
            for cell in non_empty
            if isinstance(cell, (int, float, Decimal, date, datetime))
            or self._looks_like_number(cell)
            or self._looks_like_date(cell)
        )
        return numeric_like >= max(2, len(non_empty) // 2)

    def _looks_like_header_continuation_row(self, row: tuple[Any, ...]) -> bool:
        non_empty = [cell for cell in row if cell not in (None, "")]
        if not non_empty:
            return False
        if any(str(cell).strip() in {"涨跌", "较昨日", "环比"} for cell in non_empty):
            return True
        zero_like = sum(1 for cell in non_empty if str(cell).strip() in {"0", "0.0", "-", "--"})
        return zero_like >= max(3, len(non_empty) // 2)

    def _looks_like_header_support_row(self, row: tuple[Any, ...]) -> bool:
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) < 2:
            return False
        textual = sum(1 for cell in non_empty if self._is_textual_header_cell(cell))
        return textual >= max(1, len(non_empty) // 2)

    def _is_textual_header_cell(self, value: Any) -> bool:
        if value in (None, ""):
            return False
        if isinstance(value, (int, float, Decimal, date, datetime)):
            return False
        text = str(value).strip()
        if not text:
            return False
        return not self._looks_like_number(text)


excel_analyzer = ExcelAnalyzer()
