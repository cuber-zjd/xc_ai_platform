import hashlib
import re
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import mammoth
from openpyxl import load_workbook
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from app.models.agent.insight import InsightCompany, InsightDataSource
from app.models.system.sys_company import SysCompany
from app.schemas.agent.insight.data_source import (
    InsightDataSourceCreate,
    InsightDataSourceFetchConfig,
    InsightDataSourceImportItem,
    InsightDataSourceImportResponse,
    InsightDataSourceUpdate,
)
from app.services.agent.insight.data_source_service import insight_data_source_service


@dataclass
class RequirementSourceDraft:
    row_no: int
    source_name: str
    source_type: str
    base_url: str | None
    keywords: list[str]
    project_name: str | None
    channel_name: str | None
    search_need: str | None
    target_fields: list[str]
    source_document: str
    company_hint: str | None = None


class InsightRequirementImportService:
    unsupported_markers = ("登录", "验证码", "付费", "小红书", "淘宝", "京东", "盒马", "亚马逊", "Lazada", "Shopee")

    async def import_files(
        self,
        db: AsyncSession,
        *,
        files: list[tuple[str, bytes]],
        user_id: int | None,
        is_admin: bool,
    ) -> InsightDataSourceImportResponse:
        drafts, unsupported = await self._parse_files(db, files=files)
        companies = list((await db.exec(select(InsightCompany).where(InsightCompany.is_deleted == 0))).all())
        sys_companies = list((await db.exec(select(SysCompany).where(SysCompany.is_deleted == 0))).all())
        response = InsightDataSourceImportResponse(file_count=len(files), parsed_count=len(drafts), unsupported_channels=unsupported)
        for draft in drafts:
            company = self._match_company(draft, companies, sys_companies)
            item = self._build_import_item(draft, company)
            try:
                payload = self._to_data_source_payload(draft, company)
                existing = await self._find_existing_data_source(db, payload.source_code or "")
                if existing:
                    if existing.is_deleted:
                        self._apply_payload_to_existing(existing, payload, user_id)
                        await db.commit()
                        await db.refresh(existing)
                        updated = await insight_data_source_service.get_data_source(
                            db,
                            existing.id or 0,
                            user_id=user_id or 0,
                            is_admin=True,
                        )
                    else:
                        updated = await insight_data_source_service.update_data_source(
                            db,
                            existing.id or 0,
                            InsightDataSourceUpdate(**payload.model_dump(exclude={"source_code"})),
                            user_id,
                            is_admin=True,
                        )
                    item.status = "updated"
                    item.data_source_id = updated.id
                    response.updated_count += 1
                else:
                    created = await insight_data_source_service.create_data_source(db, payload, user_id)
                    item.status = "created"
                    item.data_source_id = created.id
                    response.created_count += 1
            except Exception as exc:
                item.status = "failed"
                item.message = str(exc)
                response.failed_count += 1
            response.items.append(item)
        response.skipped_count = len(unsupported)
        return response

    async def preview_files(
        self,
        db: AsyncSession,
        *,
        files: list[tuple[str, bytes]],
        user_id: int | None,
        is_admin: bool,
    ) -> InsightDataSourceImportResponse:
        _ = (user_id, is_admin)
        drafts, unsupported = await self._parse_files(db, files=files)
        companies = list((await db.exec(select(InsightCompany).where(InsightCompany.is_deleted == 0))).all())
        sys_companies = list((await db.exec(select(SysCompany).where(SysCompany.is_deleted == 0))).all())
        response = InsightDataSourceImportResponse(file_count=len(files), parsed_count=len(drafts), unsupported_channels=unsupported)
        for draft in drafts:
            company = self._match_company(draft, companies, sys_companies)
            try:
                payload = self._to_data_source_payload(draft, company)
                existing = await self._find_existing_data_source(db, payload.source_code or "")
                item = self._build_import_item(draft, company)
                if existing and existing.is_deleted:
                    item.status = "will_restore"
                    item.data_source_id = existing.id
                    response.updated_count += 1
                elif existing:
                    item.status = "will_update"
                    item.data_source_id = existing.id
                    response.updated_count += 1
                else:
                    item.status = "will_create"
                    response.created_count += 1
            except Exception as exc:
                item.status = "failed"
                item.message = str(exc)
                response.failed_count += 1
            response.items.append(item)
        response.skipped_count = len(unsupported)
        return response

    async def _parse_files(
        self,
        db: AsyncSession,
        *,
        files: list[tuple[str, bytes]],
    ) -> tuple[list[RequirementSourceDraft], list[dict[str, Any]]]:
        _ = db
        drafts: list[RequirementSourceDraft] = []
        unsupported: list[dict[str, Any]] = []
        for file_name, file_bytes in files:
            suffix = Path(file_name).suffix.lower()
            if suffix in {".xlsx", ".xlsm"}:
                parsed, skipped = self._parse_xlsx(file_name, file_bytes)
            elif suffix == ".docx":
                parsed, skipped = self._parse_docx(file_name, file_bytes)
            else:
                parsed, skipped = [], [{"channel": file_name, "reason": "仅支持 .xlsx/.xlsm/.docx 文件"}]
            drafts.extend(parsed)
            unsupported.extend(skipped)
        return drafts, unsupported

    def _build_import_item(self, draft: RequirementSourceDraft, company: InsightCompany | None) -> InsightDataSourceImportItem:
        return InsightDataSourceImportItem(
            row_no=draft.row_no,
            source_name=draft.source_name,
            source_type=draft.source_type,
            base_url=draft.base_url,
            company_id=company.id if company else None,
            company_name=company.name if company else None,
            keywords=draft.keywords,
            project_name=draft.project_name,
            channel_name=draft.channel_name,
            source_document=draft.source_document,
        )

    def _parse_xlsx(self, file_name: str, file_bytes: bytes) -> tuple[list[RequirementSourceDraft], list[dict[str, Any]]]:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        drafts: list[RequirementSourceDraft] = []
        unsupported: list[dict[str, Any]] = []
        row_no = 1
        for sheet in workbook.worksheets:
            rows = [[self._cell_text(cell.value) for cell in row] for row in sheet.iter_rows()]
            header_index = next((idx for idx, row in enumerate(rows) if sum(1 for cell in row if cell) >= 2), None)
            if header_index is None:
                continue
            headers = rows[header_index]
            for raw_row in rows[header_index + 1 :]:
                values = {headers[idx]: raw_row[idx] for idx in range(min(len(headers), len(raw_row))) if headers[idx]}
                if not any(values.values()):
                    continue
                row_no += 1
                draft = self._draft_from_mapping(file_name, sheet.title, row_no, values)
                if draft:
                    drafts.append(draft)
                else:
                    unsupported.append({"channel": sheet.title, "row_no": row_no, "reason": "未识别到可导入的数据源字段"})
        return drafts, unsupported

    def _parse_docx(self, file_name: str, file_bytes: bytes) -> tuple[list[RequirementSourceDraft], list[dict[str, Any]]]:
        if "网址整理" not in Path(file_name).name:
            return [], [{"channel": file_name, "reason": "该 Word 文档作为需求说明参与人工理解，不直接拆分为数据源"}]
        text = mammoth.extract_raw_text(BytesIO(file_bytes)).value
        drafts: list[RequirementSourceDraft] = []
        unsupported: list[dict[str, Any]] = []
        for row_no, line in enumerate([item.strip() for item in text.splitlines() if item.strip()], start=1):
            urls = re.findall(r"https?://[^\s，,；;）)]+", line)
            source_type = self._classify_source_type(line, bool(urls))
            names = [part.strip(" ：:\t") for part in re.split(r"[\t，,；;、]", line) if part.strip(" ：:\t")]
            if urls:
                name = self._strip_url(names[0]) if names else urls[0]
                if not self._valid_source_name(name):
                    continue
                drafts.append(
                    RequirementSourceDraft(
                        row_no=row_no,
                        source_name=name[:180],
                        source_type=source_type,
                        base_url=urls[0],
                        keywords=self._keywords_from_text(line),
                        project_name=None,
                        channel_name=name[:100],
                        search_need=line[:500],
                        target_fields=[],
                        source_document=file_name,
                        company_hint=line,
                    )
                )
                continue
            if self._looks_like_channel_line(line):
                prefix, _, tail = line.partition("：")
                candidate_text = tail or line
                if not tail and len(line) > 80:
                    continue
                names = [part.strip(" ：:\t（）()") for part in re.split(r"[\t，,；;、]", candidate_text) if part.strip(" ：:\t（）()")]
                for name in names[:40]:
                    name = self._cleanup_source_name(name)
                    if not self._valid_source_name(name) or name in {"企业官网", "政府网", "其他网址", "御馨", "健源"}:
                        continue
                    maybe_type = self._classify_source_type(f"{prefix} {name}", False)
                    if any(marker.lower() in name.lower() for marker in self.unsupported_markers):
                        unsupported.append({"channel": name, "reason": "第一版仅采集公开可访问内容，登录态/强反爬渠道待人工接入"})
                        continue
                    drafts.append(
                        RequirementSourceDraft(
                            row_no=row_no,
                            source_name=name[:180],
                            source_type=maybe_type,
                            base_url=None,
                            keywords=self._keywords_from_text(name),
                            project_name=None,
                            channel_name=name[:100],
                            search_need=line[:500],
                            target_fields=[],
                            source_document=file_name,
                            company_hint=line,
                        )
                    )
        return drafts, unsupported

    def _draft_from_mapping(
        self,
        file_name: str,
        sheet_name: str,
        row_no: int,
        values: dict[str, str],
    ) -> RequirementSourceDraft | None:
        text = " ".join(str(value) for value in values.values() if value)
        project_name = self._pick(values, "课题/项目", "项目", "课题")
        channel_name = self._pick(values, "来源", "公众号", "名称", "网站", "平台", "渠道", "官网")
        url = self._first_url(text) or self._pick(values, "网址", "URL", "链接", "官网")
        search_need = self._pick(values, "搜集需求", "获取内容", "需求", "内容")
        target_fields = self._split_words(self._pick(values, "获取内容", "目标字段", "字段") or "")
        keywords = self._dedupe(
            self._split_words(self._pick(values, "关键词", "关键字") or "")
            + self._keywords_from_text(" ".join([project_name or "", channel_name or "", search_need or ""]))
        )
        source_name = channel_name or project_name or (url or "").replace("https://", "").replace("http://", "").split("/")[0]
        if not source_name:
            return None
        source_type = self._classify_source_type(f"{sheet_name} {source_name} {text}", bool(url))
        if source_type == "official_site" and not url:
            source_type = "multi_news"
        company_hint = " ".join([project_name or "", source_name, text])
        return RequirementSourceDraft(
            row_no=row_no,
            source_name=source_name[:180],
            source_type=source_type,
            base_url=url,
            keywords=keywords[:8] or [source_name],
            project_name=project_name,
            channel_name=channel_name,
            search_need=search_need,
            target_fields=target_fields,
            source_document=file_name,
            company_hint=company_hint,
        )

    def _to_data_source_payload(self, draft: RequirementSourceDraft, company: InsightCompany | None) -> InsightDataSourceCreate:
        source_code = "req_" + hashlib.sha1(f"{draft.source_document}|{draft.source_name}|{draft.source_type}".encode("utf-8")).hexdigest()[:20]
        filter_prompt = (
            "只保留与研发营销市场洞察相关的公开信息，重点判断其对客户需求、竞对动态、产品创新、"
            "政策法规、价格行情、专利技术或渠道趋势是否有价值。"
        )
        fetch_config = InsightDataSourceFetchConfig(
            keywords=draft.keywords[:8] or [draft.source_name],
            include_keywords=[],
            exclude_keywords=["招聘", "股票讨论灌水", "广告招商"],
            max_results=12,
            crawl_top_n=8,
            freshness="noLimit",
            enable_llm_filter=True,
            filter_prompt=filter_prompt,
            llm_min_score=0.55,
            auto_review_mode="high_confidence",
            auto_review_min_confidence=0.72,
            auto_add_to_report_pool=True,
            auto_report_folder="期初真实采集素材",
            extra={
                "project_name": draft.project_name,
                "channel_name": draft.channel_name,
                "platforms": [draft.channel_name] if draft.channel_name else [],
                "search_need": draft.search_need,
                "target_fields": draft.target_fields,
                "source_document": draft.source_document,
                "company_hint": draft.company_hint,
            },
        )
        return InsightDataSourceCreate(
            source_code=source_code,
            source_name=draft.source_name,
            source_type=draft.source_type,
            base_url=draft.base_url,
            company_id=company.id if company else None,
            fetch_frequency="daily",
            fetch_config=fetch_config,
            schedule_enabled=True,
            visibility_scope="assigned",
            status="enabled",
        )

    async def _find_existing_data_source(self, db: AsyncSession, source_code: str) -> InsightDataSource | None:
        if not source_code:
            return None
        return (await db.exec(select(InsightDataSource).where(InsightDataSource.source_code == source_code))).first()

    def _apply_payload_to_existing(self, row: InsightDataSource, payload: InsightDataSourceCreate, user_id: int | None) -> None:
        row.is_deleted = 0
        row.source_name = payload.source_name
        row.source_type = payload.source_type
        row.base_url = payload.base_url
        row.company_id = payload.company_id
        row.fetch_frequency = payload.fetch_frequency
        row.fetch_config = insight_data_source_service._normalize_fetch_config(payload.fetch_config)  # noqa: SLF001
        row.auth_config_ref = payload.auth_config_ref
        row.schedule_enabled = insight_data_source_service._resolve_schedule_enabled(payload.fetch_frequency, payload.schedule_enabled)  # noqa: SLF001
        row.next_run_time = (
            insight_data_source_service._calculate_next_run_time(payload.fetch_frequency, row.fetch_config, datetime.now())  # noqa: SLF001
            if row.schedule_enabled
            else None
        )
        row.last_schedule_status = "waiting" if row.schedule_enabled else None
        row.owner_user_id = row.owner_user_id or user_id
        row.visibility_scope = payload.visibility_scope
        row.status = payload.status
        row.update_by = str(user_id) if user_id else None

    def _match_company(
        self,
        draft: RequirementSourceDraft,
        companies: list[InsightCompany],
        sys_companies: list[SysCompany],
    ) -> InsightCompany | None:
        text = " ".join([draft.source_name, draft.company_hint or "", draft.project_name or ""])
        if "健源" in text:
            sys_company = self._find_sys_company(sys_companies, "山东香驰健源生物科技有限公司")
            return self._find_insight_company_by_sys(companies, sys_company) or self._find_insight_company(companies, "健源")
        if "御馨" in text:
            sys_company = self._find_sys_company(sys_companies, "山东御馨生物科技股份有限公司")
            return self._find_insight_company_by_sys(companies, sys_company) or self._find_insight_company(companies, "御馨")
        return self._find_insight_company(companies, draft.source_name)

    def _find_sys_company(self, sys_companies: list[SysCompany], name: str) -> SysCompany | None:
        return next((item for item in sys_companies if item.name == name or name in item.name or item.name in name), None)

    def _find_insight_company_by_sys(self, companies: list[InsightCompany], sys_company: SysCompany | None) -> InsightCompany | None:
        if not sys_company or not sys_company.id:
            return None
        return next((item for item in companies if item.sys_company_id == sys_company.id), None)

    def _find_insight_company(self, companies: list[InsightCompany], text: str) -> InsightCompany | None:
        normalized = self._normalize_name(text)
        for company in companies:
            names = [company.name, company.short_name or ""]
            if any(name and (self._normalize_name(name) in normalized or normalized in self._normalize_name(name)) for name in names):
                return company
        return None

    def _classify_source_type(self, text: str, has_url: bool) -> str:
        lower = text.lower()
        if "公众号" in text or "微信" in text:
            return "wechat_public_account"
        if any(word.lower() in lower for word in ["京东", "淘宝", "天猫", "盒马", "亚马逊", "lazada", "shopee", "电商"]):
            return "ecommerce_search"
        if "专利" in text or "wipo" in lower or "cnipa" in lower:
            return "patent_search"
        if any(word in text for word in ["政府", "部委", "政策", "法规", "市场监督", "发改委", "工信部"]):
            return "government_policy"
        if any(word in text for word in ["财经", "证券", "交易所", "东方财富", "同花顺", "雪球", "财报", "招股"]):
            return "finance_news"
        if any(word in text for word in ["Food", "食品", "饮料", "烘焙", "植物基", "营养", "产业", "行业媒体"]):
            return "industry_media"
        if has_url:
            return "official_site"
        return "multi_news"

    def _looks_like_channel_line(self, line: str) -> bool:
        if any(word in line for word in ["痛点", "需求梳理", "调研记录", "获取主要依赖", "存在以下", "平台建设", "章节"]):
            return False
        return any(word in line for word in ["企业官网", "政府网", "其他网址", "公众号", "电商", "财经", "专利", "食品", "饮料", "客户", "竞对"])

    def _cleanup_source_name(self, name: str) -> str:
        name = re.sub(r"^[A-Za-z0-9. ]+", "", name).strip()
        name = re.sub(r"（[^）]{0,20}家）", "", name)
        name = re.sub(r"^[^：:]{1,12}[：:]", "", name)
        return name.strip(" ：:\t（）()。.")

    def _valid_source_name(self, name: str) -> bool:
        if not (2 <= len(name) <= 40):
            return False
        invalid_words = ["目前", "存在", "根据", "结合", "核心", "维度", "获取", "缺乏", "不同渠道", "信息散落", "人工操作", "如下", "平台"]
        if any(word in name for word in invalid_words):
            return False
        return True

    def _pick(self, values: dict[str, str], *keys: str) -> str | None:
        for wanted in keys:
            for key, value in values.items():
                if wanted.lower() in str(key).lower() and value:
                    return str(value).strip()
        return None

    def _cell_text(self, value: object) -> str:
        return "" if value is None else str(value).strip()

    def _first_url(self, text: str) -> str | None:
        urls = re.findall(r"https?://[^\s，,；;）)]+", text)
        return urls[0] if urls else None

    def _strip_url(self, text: str) -> str:
        return re.sub(r"https?://\S+", "", text).strip(" ：:\t") or text

    def _keywords_from_text(self, text: str) -> list[str]:
        words = self._split_words(text)
        return [word for word in words if 2 <= len(word) <= 40][:10]

    def _split_words(self, text: str) -> list[str]:
        return self._dedupe([item.strip() for item in re.split(r"[\n\r,，;；、|/]+", text) if item.strip()])

    def _dedupe(self, values: list[str]) -> list[str]:
        result: list[str] = []
        seen: set[str] = set()
        for value in values:
            normalized = value.lower()
            if normalized and normalized not in seen:
                seen.add(normalized)
                result.append(value)
        return result

    def _normalize_name(self, value: str) -> str:
        return re.sub(r"(股份)?有限公司|集团|公司|inc\.?|limited|ltd\.?|co\.?", "", value.lower()).strip()


insight_requirement_import_service = InsightRequirementImportService()
