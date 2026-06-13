from collections import Counter
from io import BytesIO
from datetime import datetime, timedelta
from typing import Any
from uuid import uuid4

from sqlalchemy import func, or_
from sqlmodel import select
from sqlmodel.ext.asyncio.session import AsyncSession

from app.models.agent.insight import (
    InsightCompany,
    InsightDataSource,
    InsightIntelligence,
    InsightIntelligenceCandidate,
    InsightIntelligenceSource,
)
from app.schemas.agent.insight.company import (
    InsightCompanyCreate,
    InsightCompanyDataSourceSummary,
    InsightCompanyDetail,
    InsightCompanyImportError,
    InsightCompanyImportResponse,
    InsightCompanyListItem,
    InsightCompanyMetric,
    InsightCompanyRead,
    InsightCompanyTagStat,
    InsightCompanyTimelineItem,
    InsightCompanyTypeSlice,
    InsightCompanyUpdate,
)
from app.schemas.page import Page
from app.services.agent.insight.permission_service import insight_permission_service

COMPANY_IMPORT_HEADER_ALIASES = {
    "company_code": {"企业编码", "公司编码", "客户编码", "编码", "company_code", "code"},
    "name": {"企业名称", "公司名称", "客户名称", "名称", "企业", "公司", "客户", "name"},
    "short_name": {"简称", "企业简称", "公司简称", "客户简称", "short_name"},
    "industry": {"行业", "所属行业", "行业分类", "industry"},
    "company_type": {"企业类型", "公司类型", "客户类型", "类型", "company_type"},
    "region": {"区域", "地区", "省市", "所在地", "region"},
    "website": {"官网", "官方网站", "网站", "网址", "website"},
    "monitor_level": {"监控级别", "监控等级", "关注级别", "monitor_level"},
    "description": {"描述", "说明", "企业描述", "备注", "description"},
}

COMPANY_IMPORT_TEMPLATE_COLUMNS = "企业名称、简称、行业、企业类型、区域、官网、监控级别、描述"
COMPANY_IMPORT_MAX_ROWS = 1000
COMPANY_IMPORT_MAX_BYTES = 5 * 1024 * 1024


class InsightCompanyService:
    async def list_companies(
        self,
        db: AsyncSession,
        *,
        page: int,
        size: int,
        keyword: str | None,
        industry: str | None,
        monitor_level: str | None,
        status: str | None,
        user_id: int,
        is_admin: bool,
    ) -> Page[InsightCompanyListItem]:
        page = max(page, 1)
        size = min(max(size, 1), 100)
        filters = [InsightCompany.is_deleted == 0]
        if keyword:
            like_keyword = f"%{keyword.strip()}%"
            filters.append(
                or_(
                    InsightCompany.name.ilike(like_keyword),
                    InsightCompany.short_name.ilike(like_keyword),
                    InsightCompany.industry.ilike(like_keyword),
                )
            )
        if industry:
            filters.append(InsightCompany.industry == industry)
        if monitor_level:
            filters.append(InsightCompany.monitor_level == monitor_level)
        if status:
            filters.append(InsightCompany.status == status)
        else:
            filters.append(InsightCompany.status == "active")
        filters.append(
            await insight_permission_service.visibility_filter_for_user(
                db,
                InsightCompany,
                target_type="company",
                user_id=user_id,
                is_admin=is_admin,
            )
        )

        total = (await db.exec(select(func.count()).select_from(InsightCompany).where(*filters))).one()
        statement = (
            select(InsightCompany)
            .where(*filters)
            .order_by(InsightCompany.update_time.desc(), InsightCompany.id.desc())
            .offset((page - 1) * size)
            .limit(size)
        )
        companies = list((await db.exec(statement)).all())
        items = [await self._to_list_item(db, company, user_id=user_id, is_admin=is_admin) for company in companies]
        return Page.create(items=items, total=total, page=page, size=size)

    async def get_company_detail(
        self,
        db: AsyncSession,
        company_id: int,
        *,
        user_id: int,
        is_admin: bool,
    ) -> InsightCompanyDetail:
        company = await self._get_company(db, company_id, user_id=user_id, is_admin=is_admin)
        metrics = await self._build_metrics(db, company_id, user_id=user_id, is_admin=is_admin)
        type_distribution = await self._build_type_distribution(db, company_id, user_id=user_id, is_admin=is_admin)
        tag_stats = await self._build_tag_stats(db, company_id, user_id=user_id, is_admin=is_admin)
        data_sources = await self._list_data_sources(db, company_id, user_id=user_id, is_admin=is_admin)
        timeline = await self._list_timeline(db, company_id, user_id=user_id, is_admin=is_admin)
        base = self._to_read(company).model_dump()
        return InsightCompanyDetail(
            **base,
            metrics=metrics,
            type_distribution=type_distribution,
            tag_stats=tag_stats,
            data_sources=data_sources,
            timeline=timeline,
        )

    async def create_company(
        self,
        db: AsyncSession,
        payload: InsightCompanyCreate,
        user_id: int | None,
    ) -> InsightCompanyRead:
        company_code = payload.company_code or f"company_{uuid4().hex[:16]}"
        existing = (await db.exec(select(InsightCompany).where(InsightCompany.company_code == company_code))).first()
        if existing:
            raise ValueError("企业编码已存在")
        company = InsightCompany(
            company_code=company_code,
            sys_company_id=payload.sys_company_id,
            name=payload.name,
            short_name=payload.short_name,
            industry=payload.industry,
            company_type=payload.company_type,
            region=payload.region,
            website=payload.website,
            logo_url=payload.logo_url,
            description=payload.description,
            monitor_level=payload.monitor_level,
            owner_user_id=payload.owner_user_id or user_id,
            profile_json=payload.profile_json,
            status=payload.status,
            create_by=str(user_id) if user_id else None,
            update_by=str(user_id) if user_id else None,
        )
        db.add(company)
        await db.commit()
        await db.refresh(company)
        return self._to_read(company)

    async def import_companies_from_excel(
        self,
        db: AsyncSession,
        *,
        file_name: str | None,
        file_bytes: bytes,
        user_id: int,
    ) -> InsightCompanyImportResponse:
        suffix = (file_name or "").rsplit(".", 1)[-1].lower()
        if suffix not in {"xlsx", "xlsm"}:
            raise ValueError("当前仅支持上传 xlsx 或 xlsm 文件")
        if len(file_bytes) > COMPANY_IMPORT_MAX_BYTES:
            raise ValueError("Excel 文件不能超过 5MB")

        rows = self._parse_company_excel(file_bytes)
        if not rows:
            raise ValueError(f"未识别到可导入数据，请确认表头包含：{COMPANY_IMPORT_TEMPLATE_COLUMNS}")

        response = InsightCompanyImportResponse(total_rows=len(rows))
        for row_no, row in rows:
            try:
                payload = self._company_payload_from_import_row(row)
                if not payload.name.strip():
                    response.skipped_count += 1
                    response.errors.append(InsightCompanyImportError(row_no=row_no, reason="企业名称不能为空"))
                    continue
                company, action = await self._upsert_import_company(db, payload, user_id)
                if action == "created":
                    response.created_count += 1
                else:
                    response.updated_count += 1
                response.companies.append(self._to_read(company))
            except ValueError as exc:
                response.skipped_count += 1
                response.errors.append(InsightCompanyImportError(row_no=row_no, reason=str(exc)))

        await db.commit()
        return response

    async def update_company(
        self,
        db: AsyncSession,
        company_id: int,
        payload: InsightCompanyUpdate,
        user_id: int | None,
        *,
        is_admin: bool = False,
    ) -> InsightCompanyRead:
        company = await self._get_company(db, company_id, user_id=user_id, is_admin=is_admin, permission="edit")
        for field, value in payload.model_dump(exclude_unset=True).items():
            setattr(company, field, value)
        company.update_by = str(user_id) if user_id else None
        company.update_time = datetime.now()
        await db.commit()
        await db.refresh(company)
        return self._to_read(company)

    async def _upsert_import_company(
        self,
        db: AsyncSession,
        payload: InsightCompanyCreate,
        user_id: int,
    ) -> tuple[InsightCompany, str]:
        company = await self._find_import_company(db, payload, user_id)
        now = datetime.now()
        if company:
            company.name = payload.name
            company.sys_company_id = payload.sys_company_id
            company.short_name = payload.short_name
            company.industry = payload.industry
            company.company_type = payload.company_type
            company.region = payload.region
            company.website = payload.website
            company.description = payload.description
            company.monitor_level = payload.monitor_level
            company.status = payload.status
            company.update_by = str(user_id)
            company.update_time = now
            return company, "updated"

        company = InsightCompany(
            company_code=payload.company_code or f"company_{uuid4().hex[:16]}",
            sys_company_id=payload.sys_company_id,
            name=payload.name,
            short_name=payload.short_name,
            industry=payload.industry,
            company_type=payload.company_type,
            region=payload.region,
            website=payload.website,
            logo_url=payload.logo_url,
            description=payload.description,
            monitor_level=payload.monitor_level,
            owner_user_id=user_id,
            profile_json=payload.profile_json,
            status=payload.status,
            create_by=str(user_id),
            update_by=str(user_id),
        )
        db.add(company)
        await db.flush()
        return company, "created"

    async def _find_import_company(
        self,
        db: AsyncSession,
        payload: InsightCompanyCreate,
        user_id: int,
    ) -> InsightCompany | None:
        if payload.company_code:
            existing = (
                await db.exec(
                    select(InsightCompany).where(
                        InsightCompany.company_code == payload.company_code,
                        InsightCompany.is_deleted == 0,
                    )
                )
            ).first()
            if existing and existing.owner_user_id not in {None, user_id}:
                raise ValueError("企业编码已被其他用户使用")
            return existing

        return (
            await db.exec(
                select(InsightCompany).where(
                    InsightCompany.name == payload.name,
                    InsightCompany.owner_user_id == user_id,
                    InsightCompany.is_deleted == 0,
                )
            )
        ).first()

    def _parse_company_excel(self, file_bytes: bytes) -> list[tuple[int, dict[str, str]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("缺少 openpyxl 依赖，无法读取 Excel 文件") from exc

        try:
            workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = workbook.active
        except Exception as exc:
            raise ValueError("Excel 文件解析失败，请确认文件未损坏") from exc

        header_map: dict[int, str] | None = None
        data: list[tuple[int, dict[str, str]]] = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [self._cell_to_text(value) for value in row]
            if header_map is None:
                current_header = self._match_import_headers(values)
                if current_header and "name" in current_header.values():
                    header_map = current_header
                continue
            if len(data) >= COMPANY_IMPORT_MAX_ROWS:
                break
            normalized = {
                field: values[column_index].strip()
                for column_index, field in header_map.items()
                if column_index < len(values) and values[column_index].strip()
            }
            if any(normalized.values()):
                data.append((row_index, normalized))
        return data

    def _match_import_headers(self, values: list[str]) -> dict[int, str] | None:
        header_map: dict[int, str] = {}
        for index, raw_value in enumerate(values):
            normalized = self._normalize_header(raw_value)
            for field, aliases in COMPANY_IMPORT_HEADER_ALIASES.items():
                if normalized in {self._normalize_header(alias) for alias in aliases}:
                    header_map[index] = field
                    break
        return header_map or None

    def _company_payload_from_import_row(self, row: dict[str, str]) -> InsightCompanyCreate:
        return InsightCompanyCreate(
            company_code=self._empty_to_none(row.get("company_code")),
            name=(row.get("name") or "").strip(),
            short_name=self._empty_to_none(row.get("short_name")),
            industry=self._empty_to_none(row.get("industry")),
            company_type=self._empty_to_none(row.get("company_type")),
            region=self._empty_to_none(row.get("region")),
            website=self._empty_to_none(row.get("website")),
            monitor_level=self._normalize_monitor_level(row.get("monitor_level")),
            description=self._empty_to_none(row.get("description")),
            status="active",
        )

    def _normalize_monitor_level(self, value: str | None) -> str:
        text = (value or "").strip().lower()
        if not text:
            return "normal"
        mapping = {
            "普通": "normal",
            "普通监控": "normal",
            "normal": "normal",
            "重点": "key",
            "重点客户": "key",
            "key": "key",
            "竞对": "competitor",
            "重点竞对": "competitor",
            "competitor": "competitor",
            "观察": "watch",
            "观察名单": "watch",
            "watch": "watch",
        }
        return mapping.get(text, text[:20])

    def _cell_to_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        return str(value).strip()

    def _empty_to_none(self, value: str | None) -> str | None:
        text = (value or "").strip()
        return text or None

    def _normalize_header(self, value: str) -> str:
        return value.strip().lower().replace(" ", "").replace("_", "").replace("-", "")

    async def _get_company(
        self,
        db: AsyncSession,
        company_id: int,
        *,
        user_id: int | None,
        is_admin: bool,
        permission: str = "view",
    ) -> InsightCompany:
        filters = [InsightCompany.id == company_id, InsightCompany.is_deleted == 0]
        filters.append(
            await insight_permission_service.visibility_filter_for_user(
                db,
                InsightCompany,
                target_type="company",
                user_id=user_id,
                is_admin=is_admin,
                permission=permission,
            )
        )
        company = (await db.exec(select(InsightCompany).where(*filters))).first()
        if not company:
            raise ValueError("企业档案不存在或无权访问")
        return company

    async def _to_list_item(
        self,
        db: AsyncSession,
        company: InsightCompany,
        *,
        user_id: int,
        is_admin: bool,
    ) -> InsightCompanyListItem:
        intelligence_count = await self._count_visible_intelligence(db, company.id, user_id=user_id, is_admin=is_admin)
        candidate_count = await self._count(db, InsightIntelligenceCandidate, company.id)
        data_source_count = await self._count_visible_data_sources(db, company.id, user_id=user_id, is_admin=is_admin)
        latest_intelligence_time = (
            await db.exec(
                select(func.max(InsightIntelligence.publish_time))
                .where(
                    InsightIntelligence.company_id == company.id,
                    InsightIntelligence.is_deleted == 0,
                    InsightIntelligence.status == "active",
                    await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()
        return InsightCompanyListItem(
            **self._to_read(company).model_dump(),
            intelligence_count=intelligence_count,
            candidate_count=candidate_count,
            data_source_count=data_source_count,
            latest_intelligence_time=latest_intelligence_time,
        )

    async def _count(self, db: AsyncSession, model, company_id: int | None) -> int:
        return (
            await db.exec(
                select(func.count())
                .select_from(model)
                .where(
                    model.company_id == company_id,
                    model.is_deleted == 0,
                )
            )
        ).one()

    async def _count_visible_intelligence(self, db: AsyncSession, company_id: int | None, *, user_id: int, is_admin: bool) -> int:
        return (
            await db.exec(
                select(func.count())
                .select_from(InsightIntelligence)
                .where(
                    InsightIntelligence.company_id == company_id,
                    InsightIntelligence.is_deleted == 0,
                    InsightIntelligence.status == "active",
                    await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()

    async def _count_visible_data_sources(self, db: AsyncSession, company_id: int | None, *, user_id: int, is_admin: bool) -> int:
        return (
            await db.exec(
                select(func.count())
                .select_from(InsightDataSource)
                .where(
                    InsightDataSource.company_id == company_id,
                    InsightDataSource.is_deleted == 0,
                    await self._data_source_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()

    async def _build_metrics(self, db: AsyncSession, company_id: int, *, user_id: int, is_admin: bool) -> list[InsightCompanyMetric]:
        now = datetime.now()
        current_cutoff = now - timedelta(days=30)
        previous_cutoff = now - timedelta(days=60)
        current_count = await self._count_intelligence_since(db, company_id, current_cutoff, user_id=user_id, is_admin=is_admin)
        previous_count = await self._count_intelligence_between(db, company_id, previous_cutoff, current_cutoff, user_id=user_id, is_admin=is_admin)
        candidate_count = await self._count(db, InsightIntelligenceCandidate, company_id)
        source_count = await self._count_visible_data_sources(db, company_id, user_id=user_id, is_admin=is_admin)
        high_count = (
            await db.exec(
                select(func.count())
                .select_from(InsightIntelligence)
                .where(
                    InsightIntelligence.company_id == company_id,
                    InsightIntelligence.is_deleted == 0,
                    InsightIntelligence.status == "active",
                    InsightIntelligence.importance_level == "high",
                    await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()
        return [
            InsightCompanyMetric(key="recent_intelligence", label="近30天情报", value=current_count, compare_label="较前30天", delta=current_count - previous_count),
            InsightCompanyMetric(key="candidate_count", label="候选情报", value=candidate_count, compare_label="待审核/已沉淀", delta=0),
            InsightCompanyMetric(key="data_sources", label="关联数据源", value=source_count, compare_label="可采集来源", delta=0),
            InsightCompanyMetric(key="high_importance", label="高关注情报", value=high_count, compare_label="正式情报", delta=0),
        ]

    async def _count_intelligence_since(self, db: AsyncSession, company_id: int, cutoff: datetime, *, user_id: int, is_admin: bool) -> int:
        return (
            await db.exec(
                select(func.count())
                .select_from(InsightIntelligence)
                .where(
                    InsightIntelligence.company_id == company_id,
                    InsightIntelligence.is_deleted == 0,
                    InsightIntelligence.status == "active",
                    InsightIntelligence.create_time >= cutoff,
                    await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()

    async def _count_intelligence_between(
        self,
        db: AsyncSession,
        company_id: int,
        start: datetime,
        end: datetime,
        *,
        user_id: int,
        is_admin: bool,
    ) -> int:
        return (
            await db.exec(
                select(func.count())
                .select_from(InsightIntelligence)
                .where(
                    InsightIntelligence.company_id == company_id,
                    InsightIntelligence.is_deleted == 0,
                    InsightIntelligence.status == "active",
                    InsightIntelligence.create_time >= start,
                    InsightIntelligence.create_time < end,
                    await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                )
            )
        ).one()

    async def _build_type_distribution(self, db: AsyncSession, company_id: int, *, user_id: int, is_admin: bool) -> list[InsightCompanyTypeSlice]:
        statement = (
            select(InsightIntelligence.intelligence_type, func.count())
            .where(
                InsightIntelligence.company_id == company_id,
                InsightIntelligence.is_deleted == 0,
                InsightIntelligence.status == "active",
                await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
            )
            .group_by(InsightIntelligence.intelligence_type)
            .order_by(func.count().desc())
        )
        rows = list((await db.exec(statement)).all())
        total = sum(row[1] for row in rows) or 1
        return [
            InsightCompanyTypeSlice(label=row[0], count=row[1], percent=round(row[1] * 100 / total, 1))
            for row in rows
        ]

    async def _build_tag_stats(self, db: AsyncSession, company_id: int, *, user_id: int, is_admin: bool) -> list[InsightCompanyTagStat]:
        counter: Counter[str] = Counter()
        intelligence_rows = list(
            (
                await db.exec(
                    select(InsightIntelligence.raw_payload)
                    .where(
                        InsightIntelligence.company_id == company_id,
                        InsightIntelligence.is_deleted == 0,
                        InsightIntelligence.status == "active",
                        await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
                    )
                    .limit(200)
                )
            ).all()
        )
        candidate_rows = list(
            (
                await db.exec(
                    select(InsightIntelligenceCandidate.suggested_tags)
                    .where(
                        InsightIntelligenceCandidate.company_id == company_id,
                        InsightIntelligenceCandidate.is_deleted == 0,
                    )
                    .limit(200)
                )
            ).all()
        )
        for payload in intelligence_rows:
            if isinstance(payload, dict):
                self._count_tags(counter, payload.get("suggested_tags"))
        for tags in candidate_rows:
            self._count_tags(counter, tags)
        return [InsightCompanyTagStat(name=name, count=count) for name, count in counter.most_common(12)]

    def _count_tags(self, counter: Counter[str], value: object) -> None:
        if not isinstance(value, list):
            return
        for item in value:
            name = item.get("name") if isinstance(item, dict) else item
            if isinstance(name, str) and name.strip():
                counter[name.strip()] += 1

    async def _list_data_sources(self, db: AsyncSession, company_id: int, *, user_id: int, is_admin: bool) -> list[InsightCompanyDataSourceSummary]:
        statement = (
            select(InsightDataSource)
            .where(
                InsightDataSource.company_id == company_id,
                InsightDataSource.is_deleted == 0,
                await self._data_source_visibility_filter(db, user_id=user_id, is_admin=is_admin),
            )
            .order_by(InsightDataSource.update_time.desc())
            .limit(20)
        )
        return [
            InsightCompanyDataSourceSummary(
                id=row.id or 0,
                source_name=row.source_name,
                source_type=row.source_type,
                status=row.status,
                last_success_time=row.last_success_time,
            )
            for row in (await db.exec(statement)).all()
        ]

    async def _list_timeline(self, db: AsyncSession, company_id: int, *, user_id: int, is_admin: bool) -> list[InsightCompanyTimelineItem]:
        statement = (
            select(InsightIntelligence)
            .where(
                InsightIntelligence.company_id == company_id,
                InsightIntelligence.is_deleted == 0,
                InsightIntelligence.status == "active",
                await self._intelligence_visibility_filter(db, user_id=user_id, is_admin=is_admin),
            )
            .order_by(InsightIntelligence.publish_time.desc().nullslast(), InsightIntelligence.create_time.desc())
            .limit(20)
        )
        intelligences = list((await db.exec(statement)).all())
        if not intelligences:
            return []
        source_statement = (
            select(InsightIntelligenceSource)
            .where(
                InsightIntelligenceSource.intelligence_id.in_([item.id for item in intelligences if item.id]),
                InsightIntelligenceSource.is_deleted == 0,
            )
            .order_by(InsightIntelligenceSource.credibility_score.desc(), InsightIntelligenceSource.create_time.desc())
        )
        sources_by_id: dict[int, InsightIntelligenceSource] = {}
        for source in (await db.exec(source_statement)).all():
            sources_by_id.setdefault(source.intelligence_id, source)
        return [
            InsightCompanyTimelineItem(
                id=item.id or 0,
                title=item.title,
                summary=item.summary,
                intelligence_type=item.intelligence_type,
                importance_level=item.importance_level,
                publish_time=item.publish_time,
                create_time=item.create_time,
                primary_source_url=sources_by_id.get(item.id or 0).source_url if sources_by_id.get(item.id or 0) else None,
                primary_source_title=sources_by_id.get(item.id or 0).source_title if sources_by_id.get(item.id or 0) else None,
            )
            for item in intelligences
        ]

    async def _intelligence_visibility_filter(self, db: AsyncSession, *, user_id: int, is_admin: bool):
        return await insight_permission_service.visibility_filter_for_user(
            db,
            InsightIntelligence,
            target_type="intelligence",
            user_id=user_id,
            is_admin=is_admin,
        )

    async def _data_source_visibility_filter(self, db: AsyncSession, *, user_id: int, is_admin: bool):
        return await insight_permission_service.visibility_filter_for_user(
            db,
            InsightDataSource,
            target_type="data_source",
            user_id=user_id,
            is_admin=is_admin,
        )

    def _to_read(self, row: InsightCompany) -> InsightCompanyRead:
        return InsightCompanyRead(
            id=row.id,
            create_time=row.create_time,
            update_time=row.update_time,
            create_by=row.create_by,
            update_by=row.update_by,
            comment=row.comment,
            is_deleted=row.is_deleted,
            company_code=row.company_code,
            sys_company_id=row.sys_company_id,
            name=row.name,
            short_name=row.short_name,
            industry=row.industry,
            company_type=row.company_type,
            region=row.region,
            website=row.website,
            logo_url=row.logo_url,
            description=row.description,
            monitor_level=row.monitor_level,
            owner_user_id=row.owner_user_id,
            profile_json=row.profile_json,
            status=row.status,
        )


insight_company_service = InsightCompanyService()
